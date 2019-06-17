# -*- coding: utf-8 -*-

from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

#from django.http import HttpResponse
#from django.contrib.auth.models import User

from django.shortcuts import render, redirect
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import UserCreationForm
#from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse 
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.contrib.auth.decorators import login_required
from empleados.models import Empleado
from empleados.metodos import set_values_emp
from candidatos.models import *
from candidatos.forms import *
from .models import *
from empleados.forms import *
from candidatos.metodos import *

from .forms import *
from datetime import datetime, timedelta
from django.db import IntegrityError
from django.views.defaults import page_not_found
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
###############

###############

############### Exportar a pdf

def mi_error_404(request):
    nombre_template = '404.html'
    return page_not_found(request, template_name=nombre_template)

def mod_fecha(fec):
    if fec not in [None,'']:
        return (datetime.strptime(fec, '%d/%m/%Y').strftime('%Y-%m-%d'))
    else:
        return fec

def enviar_email(request):
    name=request.POST.get('nombre')
    email=request.POST.get('correo')
    destino=request.POST.get('correo_cand')
    message=request.POST.get('mensaje')
    print(destino)
    body=render_to_string(
        'email_content.html',{
            'name':name,
            'email':email,
            'message':message,
        },
    )
    email_message = EmailMessage(
        subject='Mensaje de usuario',
        body=body,
        from_email=email,
        #to=['juvenciolugo@hotmail.com'],
        to=[destino],
    )
    email_message.content_subtype = 'html'
    email_message.send()
    return redirect('candLst')

def enviar2_email(request):
    name=request.POST.get('nombre')
    email=request.POST.get('correo')
    destino=request.POST.get('correo_emp')
    print(destino)
    message=request.POST.get('mensaje')
    body=render_to_string(
        'email_content.html',{
            'name':name,
            'email':email,
            'message':message,
        },
    )
    email_message = EmailMessage(
        subject='Mensaje de usuario',
        body=body,
        from_email=email,
        #to=['juvenciolugo@gmail.com.com'],
        to=[destino],
    )
    email_message.content_subtype = 'html'
    email_message.send()
    #aviso de envio
    return redirect('empLst')

class ReportePersonalizadoPDF(TemplateView):
    def get(self, request,*args,**kwargs):
        import io  
        from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle  
        from reportlab.lib.styles import getSampleStyleSheet  
        from reportlab.lib import colors  
        from reportlab.lib.pagesizes import letter ,A4
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import Table 
        from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER #, TA_RIGHT
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase.pdfmetrics import stringWidth
       
        
        styles = getSampleStyleSheet()  

        #styleR = styles["Normal"]
        #styleR.alignment = TA_RIGHT

        styleN = styles["Normal"]
        styleN.alignment = TA_LEFT
        
        
        styleC = styles["Normal"]
        styleC.alignment = TA_CENTER

        pk=kwargs['id']
        cand =  Candidato.objects.get(id=pk)

        response = HttpResponse(content_type='application/pdf')  
        buff = io.BytesIO()  
        c = canvas.Canvas(buff, pagesize=A4)
        doc = SimpleDocTemplate(buff,  
               pagesize=letter,  
               rightMargin=40,  
               leftMargin=40,  
               topMargin=60,  
               bottomMargin=18,  
               )  
        candidatos = []  
        
        #ejemplo = Paragraph("Mitexto", styleN)  
        #candidatos.append(ejemplo) 
        c.drawImage("static/img/logoNuevo.png", 200, 770,190,60)
        c.setLineWidth(.3)
        c.setFont('Helvetica', 22)
        c.drawString(200, 750, 'Solicitud de empleo')
        c.setFont('Helvetica', 12)
        #c.drawString(30, 735, 'Report')
        
        

        c.setFont('Helvetica', 6)
        c.drawString(400, 730,'    Fecha           Fuente de reclutamiento')
        c.drawString(400, 720,cand.fecha_solicitud.strftime('%d/%m/%Y'))
        c.drawString(450, 720,cand.fuente_recluta)
        c.drawString(350, 710,'Puesto Solicitado                            Sueldo deseado')
        c.drawString(350, 700,cand.puesto_solicitado)
        c.drawString(450, 700,cand.sueldo_deseado)
        c.drawString(50, 690,'Apellido Paterno                    Apellido Materno                    Nombre(s)                    Sexo        Estado Civil')
        c.drawString(50, 680,cand.apellido_paterno+' '+cand.apellido_materno+' '+cand.nombre)
        c.drawString(260, 680,cand.sexo)
        c.drawString(295, 680,cand.estado_civil)
        c.drawString(50, 670,'Edad  Fecha de nacimiento   Lugar de Nacimiento                  Teléfono')
        c.drawString(50, 660,cand.edad)
        c.drawString(80, 660,str(cand.fecha_nac.strftime('%d/%m/%Y')))
        c.drawString(130, 660,cand.lugar_nac)
        c.drawString(213, 660,str(cand.tel))
        c.drawString(50, 650,'Calle No                                                           Colonia                                    Estado                  CP      Tiempo de trayetoria de su casa')
        c.drawString(50, 640,cand.calle)
        c.drawString(173, 640,cand.colonia)
        if cand.esdo not in[None,'']:
            c.drawString(250, 640,cand.esdo)
        else:
            c.drawString(250, 640,'')

        c.drawString(300, 640,str(cand.cp))
        c.drawString(350, 640,str(cand.trayectoria_de_casa))
        c.drawString(50, 630,'R.F.C.             Afiliación al IMSS      No. de cartilla Militar         Tipo y No de licencia de manejo')
        c.drawString(50, 620,str(cand.rfc))
        c.drawString(100, 620,str(cand.imss))
        c.drawString(150, 620,str(cand.cartilla))
        c.drawString(230, 620,str(cand.tipo_licencia))
        c.drawString(270, 620,str(cand.licencia))
        
        
        styles = getSampleStyleSheet()
        styleBH = styles["BodyText"]
        styleBH.alignment = TA_CENTER
        styleBH.fontSize = 6

        nivel = Paragraph('''NIVEL ESCOLAR''',styleBH)
        institucion = Paragraph('''INSTITUCIÓN''',styleBH)
        annios = Paragraph('''AÑOS''',styleBH)
        inicio = Paragraph('''INICIO''',styleBH)
        termino = Paragraph('''TERMINO''',styleBH)
        documento = Paragraph('''DOCUMENTO''',styleBH)
        
        data = []
        data.append([nivel,institucion,annios,inicio,termino,documento])
        
        nivel =Paragraph('Primaria',styleBH)
        institucion =Paragraph(str(cand.primaria),styleBH)
        annios =Paragraph(str(cand.primaria_annios),styleBH)
        inicio =Paragraph(str(cand.primaria_inicio),styleBH)
        termino =Paragraph(str(cand.primaria_termino),styleBH)
        documento=Paragraph(str(cand.primaria_documento),styleBH)
        data.append([nivel,institucion,annios,inicio,termino,documento])

        nivel =Paragraph('Secundaria',styleBH)
        institucion =Paragraph(str(cand.secundaria),styleBH)
        annios =Paragraph(str(cand.secundaria_annios),styleBH)
        inicio =Paragraph(str(cand.secundaria_inicio),styleBH)
        termino =Paragraph(str(cand.secundaria_termino),styleBH)
        documento=Paragraph(str(cand.secundaria_documento),styleBH)
        data.append([nivel,institucion,annios,inicio,termino,documento])
        
        nivel =Paragraph('Preparatoria',styleBH)
        institucion =Paragraph(str(cand.preparatoria),styleBH)
        annios =Paragraph(str(cand.preparatoria_annios),styleBH)
        inicio =Paragraph(str(cand.preparatoria_inicio),styleBH)
        termino =Paragraph(str(cand.preparatoria_termino),styleBH)
        documento=Paragraph(str(cand.preparatoria_documento),styleBH)
        data.append([nivel,institucion,annios,inicio,termino,documento])

        nivel =Paragraph('Tecnica',styleBH)
        institucion =Paragraph(str(cand.tecnica),styleBH)
        annios =Paragraph(str(cand.tecnica_annios),styleBH)
        inicio =Paragraph(str(cand.tecnica_inicio),styleBH)
        termino =Paragraph(str(cand.tecnica_termino),styleBH)
        documento=Paragraph(str(cand.tecnica_documento),styleBH)
        data.append([nivel,institucion,annios,inicio,termino,documento])


        
        #datos=(cand.fecha_solicitud.strftime('%d-%m-%Y'), cand.puesto_solicitado)
        #t= Table([headings]+[datos])
        #t.setStyle([('GRID', (0, 0), (1, -1), 1, colors.dodgerblue)])
        #candidatos.append(t) 
        
        #table size
        width, height = A4
        high = 500
        table = Table(data, colWidths=[3.5 * cm, 6 * cm, 1.3 * cm,1.4*cm,2*cm,3*cm])
        table.setStyle(TableStyle([
            ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
            ('BOX',(0,0),(-1,-1), 0.25,colors.black),
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, 30, high)
        
        stylePro = styles["BodyText"]
        stylePro.alignment = TA_CENTER
        stylePro.fontSize = 6

        

        #contenido de tabla profesional
        estudios=Estudios_pro.objects.filter(candidato=cand)
        high = 446
        
        for est in estudios:
            data = []
            tipo = Paragraph(str(est.estudios_tipo),stylePro)
            institucion = Paragraph(str(est.estudios_escuela),stylePro)
            annios = Paragraph(str(est.estudios_annios),stylePro)
            inicio = Paragraph(str(est.estudios_inicio),stylePro)
            termino = Paragraph(str(est.estudios_termino),stylePro)
            documento = Paragraph(str(est.estudios_documento),stylePro)
            data.append([tipo,institucion,annios,inicio,termino,documento])

            carrera = Paragraph('Estudio',stylePro)
            tesis=Paragraph('Tésis',stylePro)
            cedula=Paragraph('Cédula',stylePro)
            data.append([carrera,'',tesis,'','',cedula])

            carrera = Paragraph(str(est.estudios_nombre),stylePro)
            tesis=Paragraph(str(est.estudios_tesis),stylePro)
            cedula=Paragraph(str(est.estudios_cedula),stylePro)
            
            data.append([carrera,'',tesis,'','',cedula])
            
        
            table = Table(data, colWidths=[3.5 * cm, 6 * cm, 1.3 * cm,1.4*cm,2*cm,3*cm])
            table.setStyle(TableStyle([
                ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
                ('BOX',(0,0),(-1,-1), 0.25,colors.black),
                ('SPAN',(0,1),(1,1)),
                ('SPAN',(0,2),(1,2)),
                ('SPAN',(2,1),(4,1)),
                ('SPAN',(2,2),(4,2)),
            
            ]))
            table.wrapOn(c, width, height)
            table.drawOn(c, 30, high)
            high-=54
        high-=10
        c.drawString(50, high,'¿Estudia actualmente? '+str(cand.estudia_actualmente)+'        ¿Que?        '+str(cand.estudia_que)+'       ¿Donde?        '+str(cand.estudia_donde))
        high-=10
        c.drawString(50, high,'Horario :   '+str(cand.estudia_horario)+'      Fecha que finaliza :     '+str(cand.estudia_termino))
        c.showPage()
        ##Hoja 2
        stylePro.alignment = TA_LEFT
        data = []
        high=800
        idiomas=Idioma_candidato.objects.filter(candidato=cand)
        idiomas_lst=''
        for idi in idiomas:
             idiomas_lst ='IDIOMA(S): '+str(idi.idioma)+' %'+str(idi.idioma_porcentaje)
        if not idiomas_lst in [None,'']:
            data.append([Paragraph(idiomas_lst,stylePro)])
            table = Table(data, colWidths=[19 * cm])
            table.setStyle(TableStyle([
                ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
                ('BOX',(0,0),(-1,-1), 0.25,colors.black),
            ]))
            table.wrapOn(c, width, height)
            table.drawOn(c, 30, high)
        high-=18
        data = []
        data.append([Paragraph('Máqunas, equipos y herramientas que puede manejar : '+str(cand.maquinas_equipos),stylePro)])
        table = Table(data, colWidths=[19 * cm])
        table.setStyle(TableStyle([
            ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
            ('BOX',(0,0),(-1,-1), 0.25,colors.black),
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, 30, high)

        #Datos personales
        
        high-=90
        c.setFont('Helvetica', 6)
        c.drawString(30, high+75,'DATOS PERSONALES')
        data = []
        des=''
        nom = Paragraph('''Nombre''',stylePro)
        edad = Paragraph('''Edad''',stylePro)
        ocu = Paragraph('''Ocupación''',stylePro)
        vive = Paragraph('''Vive''',stylePro)
        data.append(['',nom,edad,ocu,vive])
        
        des= Paragraph('Padre',stylePro)
        nom = Paragraph(str(cand.padre_nombre),stylePro)
        edad = Paragraph(str(cand.padre_edad),stylePro)
        ocu = Paragraph(str(cand.padre_ocupacion),stylePro)
        vive = Paragraph(str(cand.padre_vive),stylePro)
        data.append([des,nom,edad,ocu,vive])
        
        des= Paragraph('Madre',stylePro)
        nom = Paragraph(str(cand.madre_nombre),stylePro)
        edad = Paragraph(str(cand.madre_edad),stylePro)
        ocu = Paragraph(str(cand.madre_ocupacion),stylePro)
        vive = Paragraph(str(cand.madre_vive),stylePro)
        data.append([des,nom,edad,ocu,vive])

        des= Paragraph('Cónyuge',stylePro)
        nom = Paragraph(str(cand.conyuge_nombre),stylePro)
        edad = Paragraph(str(cand.conyuge_edad),stylePro)
        ocu = Paragraph(str(cand.conyuge_ocupacion),stylePro)
        vive = Paragraph(str(cand.conyuge_vive),stylePro)
        data.append([des,nom,edad,ocu,vive])
        table = Table(data, colWidths=[1.6*cm,6 * cm, 1 * cm, 6 * cm,1.4*cm])
        table.setStyle(TableStyle([
            ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
            ('BOX',(0,0),(-1,-1), 0.25,colors.black),
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, 30, high)
        ###HERMANOS
        hermanos=Hermano_candidato.objects.filter(candidato=cand)
        
        nher=0
        for herm in hermanos:
            nher+=1
            high -= 18
            data = []
            des = Paragraph('Hermano',stylePro)
            nom = Paragraph(str(herm.hermano_nombre),stylePro)
            edad = Paragraph(str(herm.hermano_edad),stylePro)
            ocu = Paragraph(str(herm.hermano_ocupacion),stylePro)
            data.append([des,nom,edad,ocu,''])
            table = Table(data, colWidths=[1.6*cm,6 * cm, 1 * cm, 6 * cm,1.4*cm])
            table.setStyle(TableStyle([
                ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
                ('BOX',(0,0),(-1,-1), 0.25,colors.black),
            ]))
            table.wrapOn(c, width, height)
            table.drawOn(c, 30, high)
            #high-=54
        ###HIJOS
        
        hijos=Hijo_candidato.objects.filter(candidato=cand)
        #if (nher==1):
        #    high += 36
        
        for hijo in hijos:
            high -= 18
            data = []
            des = Paragraph('Hijo',stylePro)
            nom = Paragraph(str(hijo.hijo_nombre),stylePro)
            edad = Paragraph(str(hijo.hijo_edad),stylePro)
            ocu = Paragraph(str(hijo.hijo_ocupacion),stylePro)
            data.append([des,nom,edad,ocu,''])
            table = Table(data, colWidths=[1.6*cm,6 * cm, 1 * cm, 6 * cm,1.4*cm])
            table.setStyle(TableStyle([
                ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
                ('BOX',(0,0),(-1,-1), 0.25,colors.black),
            ]))
            table.wrapOn(c, width, height)
            table.drawOn(c, 30, high)
            
        high-=18
        ##EXPERIENCIA LABORAL
        c.drawString(30, high,'EXPERIENCIA LABORAL')
        high-=15
        experiencias=Experiencia.objects.filter(candidato=cand)
        for exp in experiencias:
            c.drawString(30, high,'Empresa                                                   Dirección                                                                        Teléfono')
            c.drawString(30, high-10,str(exp.empresa_nombre))
            c.drawString(130, high-10,str(exp.empresa_direccion))
            c.drawString(280, high-10,str(exp.empresa_tel))
            high-=20
            c.drawString(30, high,'Giro                                                   Nombre del jefe inmediato                                                  Puesto del jefe inmediato')
            c.drawString(30, high-10,str(exp.empresa_giro))
            c.drawString(130, high-10,str(exp.empresa_nombre_jefe))
            c.drawString(280, high-10,str(exp.empresa_jefe_puesto))
            high-=20
            c.drawString(30, high,'Fecha de ingreso                   Salario inicial              Fecha de separación     Salario final')
            c.drawString(30, high-10,exp.empresa_fecha_ingreso.strftime('%d/%m/%Y'))
            c.drawString(120, high-10,str(exp.empresa_salario_inicio))
            fs=exp.empresa_fecha_separacion
            if fs not in [None,'']:
                c.drawString(180, high-10,exp.empresa_fecha_separacion.strftime('%d/%m/%Y'))
            else:
                c.drawString(180, high-10,'')

            c.drawString(230, high-10,str(exp.empresa_salario_final))
            high-=20
            c.drawString(30, high,'Último puesto desempeñado                   Tiempo             Departamento                    Puesto anterior                 Tiempo')
            c.drawString(30, high-10,str(exp.empresa_puesto_ultimo))
            c.drawString(120, high-10,str(exp.empresa_puesto_ultimo_tiempo))
            c.drawString(180, high-10,str(exp.empresa_puesto_ultimo_depto))
            high-=20
            c.drawString(30, high,'Puesto anterior                          Tiempo             Departamento                    Puesto anterior                 Tiempo')
            c.drawString(30, high-10,str(exp.empresa_puesto_anterior))
            c.drawString(120, high-10,str(exp.empresa_puesto_anterior_tiempo))
            c.drawString(180, high-10,str(exp.empresa_puesto_anterior_depto))

            high-=20
            c.drawString(30, high,'Experiencia en supervision : '+str(exp.experiencia_supervision)+'  No. Personas que superviso : '+str(exp.experiencia_supervision_num))
            c.drawString(30, high-10,'Motivo de la separación')
            c.drawString(30, high-20,'Mejorar el ingreso')
            high-=100

        ##DATOS GENERALES
        #' ¿Cuál? : '+str(cand.sindicato_nombre)+
        #  ' Cargo: '+str(cand.sindicato_cargo))
        high+=50
        c.drawString(30, high,'DATOS GENERALES')
        c.drawString(30, high-10,'Vive en casa: '+str(cand.vivienda_propia)+' ¿Tiene crédito infonavit? : '+str(cand.credito_infonavit)+'Pago mensual: '+str(cand.pago_infonavit))
        c.drawString(30, high-20,'¿Tiene auto propio?: '+str(cand.auto_propio)+' Marca : '+str(cand.auto_marca)+' Modelo: '+str(cand.auto_modelo)+' Seguro de vida: '+str(cand.seguro_vida)+' Monto '+ str(cand.seguro_monto))
        c.drawString(30, high-30,'¿Afiliado a un sindicato?: '+str(cand.afiliado_sindicato)+' ¿Cuál: '+str(cand.sindicato_nombre)+' Cargo: '+str(cand.sindicato_cargo))
        c.drawString(30, high-40,'¿A qué dedica su tiempo libre?: '+str(cand.tiempo_libre)+' ¿Embarazo? : '+str(cand.embarazo)+' Religión: '+str(cand.religion)+' Dispuesto a rolar turno'+ str(cand.disposicion_rolar))
        c.drawString(30, high-50,'¿Estado de salud?: '+str(cand.estado_salud)+' ¿Fuma? : '+str(cand.fuma)+' ¿Bebe?: '+str(cand.bebe)+' ¿Tatuajes?: '+str(cand.tatuajes)+' ¿Perforaciones?: '+ str(cand.perforaciones))
        c.drawString(30, high-60,'¿Dispuesto a viajar?: '+str(cand.disposicion_viajar)+' ¿Tiene ingresos extras? : '+str(cand.ingreso_extra)+' Monto?: '+str(cand.ingreso_monto)+' ¿Fuente?: '+ str(cand.ingreso_fuente))
        c.drawString(30, high-70,'¿Trabaja con nosotros algún pariente o amigo?: '+str(cand.labora_conocido)+' ¿Nombre? : '+str(cand.conocido_nombre)+' ¿Departamento?: '+ str(cand.conocido_depto))
        c.drawString(30, high-80,'En caso de ser aceptado ¿En qué fecha estaría disponible para trabajar?: '+str(cand.fecha_disponible))

        ##REFERENCIAS
        referencias=Referencia.objects.filter(candidato=cand)
        high -= 100
        stylePro.alignment = TA_CENTER
        c.drawString(30, high,'REFERENCIAS')
        high -= 20
        data = []
        nom = Paragraph('Nombre',stylePro)
        ocu = Paragraph('Ocupación',stylePro)
        tel = Paragraph('Teléfono',stylePro)
        annios = Paragraph('Años de conocerlo',stylePro)
        data.append([nom,ocu,tel,annios])
        table = Table(data, colWidths=[6 * cm, 6 * cm, 2.5 * cm,2.5*cm])
        table.setStyle(TableStyle([
              ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
              ('BOX',(0,0),(-1,-1), 0.25,colors.black),
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, 30, high)
        
        for ref in referencias:
            high -= 18
            data = []
            nom = Paragraph(str(ref.referencia_nombre),stylePro)
            ocu = Paragraph(str(ref.referencia_ocupacion),stylePro)
            tel = Paragraph(str(ref.referencia_tel),stylePro)
            annios = Paragraph(str(ref.referencia_annios_conocer),stylePro)
            
            data.append([nom,ocu,tel,annios])
            table = Table(data, colWidths=[6 * cm, 6 * cm, 2.5 * cm,2.5*cm])
            table.setStyle(TableStyle([
                ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
                ('BOX',(0,0),(-1,-1), 0.25,colors.black),
            ]))
            table.wrapOn(c, width, height)
            table.drawOn(c, 30, high)
            #high-=54
        
        
        #Firma
        high-=80
        data = []
        texto=Paragraph("CORPORATIVO AXXUM, S.A. DE C.V. LE AGRADECE LA INFORMACIÓN PROPORCIONADA, SI ALGUNO DE LOS DATOS NO FUERA VERÍDICO, TODO CONVENIO CELEBRADO CON LA EMPRESA SERIA NULO.",stylePro)
        firma=Paragraph('Firma',styleC)
        data.append([texto,firma])
        table = Table(data, colWidths=[6 * cm, 6*cm])
        table.setStyle(TableStyle([
            ('INNERGRID',(0,0),(-1,-1), 0.25,colors.black),
            ('BOX',(0,0),(-1,-1), 0.25,colors.black),
            
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, 30, high)
        c.line(235, high+17, 335, high+17)
        
        c.showPage()
        high=450
        
        

        
        styleAviT = styles["Normal"]
        styleAviT.alignment = TA_JUSTIFY
        styleAviT.fontSize = 6
        data = []
        tit_aviso=Paragraph("AVISO DE PRIVACIDAD",styleC)
        data.append([tit_aviso])
        aviso=Paragraph("     CORPORATIVO AXXUM, S.A. DE C.V., con  domicilio en la calle Rio Manzanares 308 Oficina 3,  Colonia: Del Valle,  San Pedro Garza García, Nuevo León, Código Postal 6622, "+
        "es responsable del uso que se le dé a sus datos personales y de su protección, los cuales serán tratados y resguardados con base en los principios de licitud, calidad,"+
        "consentimiento, información,  finalidad, lealtad, proporcionalidad y responsabilidad, consagrados en la Ley Federal de Protección de Datos Personales en Posesión de los Particulares."+
        "Su información personal será utilizada con el fin de encontrar al candidato ideal que pueda cubrir la vacante publicada en los distintos medios electrónicos y/o escritos. Para las finalidades antes mencionadas requerimos obtener los siguientes datos:"+
        "Datos personales : * Nombre Completo * Lugar y Fecha de Nacimiento * Sexo * Domicilio, Correo Electronico * RFC, CURP, No. IMSS, Tipo y No. De Lic. De Manejo * Estado Civil Antecedentes Academicos: * Nivel Escolar "+
        "                   * Institución y Domicilio * Años Cursados, Fecha de Ingreso y Egreso * Documento Recibido Conocimientos Generales: * Idiomas * Maquinas de Oficina o Equipo de Trabajo que Maneje * Funciones de Oficina u otros que domine."+
        "Datos Familiares: * Nombre de: Padre, Madre, Esposo, Hijos y Hermanos * Edad, Ocupación, Empresa, Dirección y Telefono de cada uno de los familiares antes mencionados Experiencia Laboral: * Empresa Actual o Ultima, Dirección y Teléfono "+
        "          * Giro de la Empresa, Nombre y Puesto del Jefe Inmediato * Fecha de  Ingreso y Sueldo, Fecha de Salida y Ultimo Sueldo * Ultimo puesto desempeñado y duración, Departamento, Puesto Anterior y duración "+
        "          * Motivo de Separación Estado de Salud y Hábitos Personales: * Tipo de casa en donde vive, Cuenta o no No con Credito INFONAVIT, Pagos Mesuales * Automóvil propio, Tipo, Modelo y Marca * Si cuenta con seguro de vida "+
        "          * Si ha sido afiliado a algun sindicato, si ha sido afianzado * A que dedica su tiempo libre * Embarazo, Religion, Rolar turnos, como es su estado de Salud, Fuma, Bebe, Enfermedades , Toma medicamentos, le han practicado alguna cirugía, etc. "+
        "          * En caso de ser contratado en que fecha puede presentarse. Referencias Personales: * Nombre, Dirección, Teléfono, Ocupación y Tiempo de conocer a las perdonas que lo refieren."+
        "Usted tiene derecho de acceder, rectificar y cancelar sus datos personales, así como de oponerse al tratamiento de los mismos o revocar el consentimiento que para tal fin nos haya otorgado, en  CORPORATIVO AXXUM, S.A. DE C.V. "+
        "En este sentido hacemos de su conocimiento que para el caso en el que  CORPORATIVO AXXUM, S.A. DE C.V., decida no contratarlo, desechará su información en un periodo de treinta días contados a partir de la fecha de firma del presente documento."+
        "En caso de querer conocer  cualquier modificación a este aviso de privacidad se  puede poner en contacto con nuestro departamento de datos personales con Victoria Hidalgo García, en  "+
        "Rio Manzanares 308 Oficina 3,  Colonia: Del Valle,  San Pedro Garza García, Nuevo León, Código Postal 6622, victoria.hidalgo@springlabs.net",styleAviT)
        data.append([aviso])
        table = Table(data, colWidths=[18 * cm])
        table.setStyle(TableStyle([
            ('INNERGRID',(0,0),(-1,-1), 0.25,colors.white),
            ('BOX',(0,0),(-1,-1), 0.25,colors.white),
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, 30, high)

        #text.setFont("Helvetica", 12)
        #text.textLine("¡Hola, mundo!")
        #text.textLine("¡Desde ReportLab y Python!")
        
        high-=300
        #text = c.beginText(50, h - 50)
        #text.setFont("Helvetica", 12)
        text ="Autorizo"
        text_width = stringWidth(text,"Helvetica",12)
        text_o = c.beginText((width-text_width)/2, high)
        text_o.textLine(text)
        c.drawText(text_o)
        text ="Nombre y firma"
        text_width = stringWidth(text,"Helvetica",12)
        text_o = c.beginText((width-text_width)/2, high-10)
        text_o.textLine(text)
        c.drawText(text_o)
        c.line(((width-text_width)/2)-10, high+15, ((width-text_width)/2) + text_width+10, high+15)

        
        c.save()    

        response.write(buff.getvalue())  
        buff.close()  
        return response


###############


class ReportePersonalizadoExcel(TemplateView):
    def get(self, request,*args,**kwargs):
        id=kwargs['id']
        

        cand = Candidato.objects.get(id=id)
        wb = Workbook()
        bandera = True
        cont = 1
        #controlador=4
        if bandera:
            ws = wb.active
            ws.title ='Hoja'+str(cont)
            bandera = False
        else:
            ws = wb.create_sheet('Hoja'+str(cont))

        #ws.title ='Hoja'+str(cont)
        #ws = wb.create_sheet('Hoja'+str(cont))
        contador=1
        for letra in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            ws[letra+str(contador)].alignment = Alignment(horizontal = "center", vertical = "center")
            ws[letra+str(contador)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top  = Side(border_style = "thin"), bottom= Side(border_style = "thin"))
            ws[letra+str(contador)].fill = PatternFill(start_color ='66FFCC',end_color='66FFCC', fill_type='solid')
            ws[letra+str(contador)].font = Font(name ='Calibri', size=12, bold=True)
            ws.column_dimensions[letra].width = 20
        
        for letra in "ABCDEFGHIJKLM":
            ws["A"+letra+str(contador)].alignment = Alignment(horizontal = "center", vertical = "center")
            ws["A"+letra+str(contador)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top  = Side(border_style = "thin"), bottom= Side(border_style = "thin"))
            ws["A"+letra+str(contador)].fill = PatternFill(start_color ='66FFCC',end_color='66FFCC', fill_type='solid')
            ws["A"+letra+str(contador)].font = Font(name ='Calibri', size=12, bold=True)
            ws.column_dimensions["A"+letra].width = 20
        ws.row_dimensions[1].height = 30

        pais= str(cand.pais_nacimiento)

        
            
        
        ws['A1'] = "NOMBRE"
        
        ws['A2'] = cand.nombre+" "+cand.apellido_paterno+" "+cand.apellido_materno
        ws['B1'] = "ZONA"
        ws['C1'] = "Puesto"
        ws['C2'] = cand.puesto_solicitado
        ws['D1'] = "Home Office"
        ws['E1'] = "Jornada"
        ws['F1'] = "Jornada Real"
        ws['G1'] = "Jornada en contrato"
        ws['H1'] = "Nacionalidad"
        ws['H2'] = pais
        ws['I1'] = "Edo civil"
        ws['I2'] = cand.estado_civil
        ws['J1'] = "Fecha de nacimiento"
        ws['J2'] = cand.fecha_nac
        ws['K1'] = "CURP"
        ws['K2'] = cand.curp
        ws['L1'] = "RFC"
        ws['L2'] = cand.rfc
        ws['M1'] = "Identificación expedida por"
        ws['N1'] = "Con número"
        ws['O1'] = "EMPRESA"
        ws['P1'] = "NSS"
        ws['P2'] = cand.imss
        ws['Q1'] = "CALLE"
        ws['Q2'] = cand.calle
        ws['R1'] = "Número exterior"
        ws['R2'] = cand.num_ext
        ws['S1'] = "Número interior"
        ws['S2'] = cand.num_int
        ws['T1'] = "COLONIA"
        ws['T2'] = cand.colonia
        ws['U1'] = "MUNICIPIO"
        ws['V1'] = "ESTADO"
        ws['V2'] = cand.esdo
        ws['W1'] = "CÓDIGO POSTAL"
        ws['W2'] = cand.cp
        ws['X1'] = "NÚMERO DE TARJETA"
        ws['Y1'] = "NO DE CUENTA"
        ws['Z1'] = "CLABE INTERBANCARIA"
        ws['AA1'] = "NOMBRE DE BANCO"
        ws['AB1'] = "Tel. casa"
        ws['AB2'] = cand.tel
        ws['AC1'] = "Tel. personal"
        ws['AD1'] = "SD IMSS"
        ws['AE1'] = "Ayuda diaria"
        ws['AF1'] = "Sueldo neto"
        ws['AG1'] = "Cantidad letra"
        ws['AH1'] = "Bono"
        ws['AI1'] = "% comisión"
        ws['AJ1'] = "Fecha de ingreso"
        ws['AK1'] = "Fecha de alta IMSS"
        ws['AL1'] = "Correo empresa"
        ws['AM1'] = "Correo personal"
        ws['AM2'] = cand.email_personal
        

        #establecer nombre de archivo
        nombre_archivo = "ReportePersonalizadoExcel.xlsx"
        #definir tipo de respuesta
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteLstPersonalizadoExcel(TemplateView):
#def ReporteLstPersonalizadoExcel(request):
    def get(self, request,*args,**kwargs):
        candidatos = Candidato.objects.all()
        wb = Workbook()
        bandera = True
        cont = 2
        
        ########
        if bandera:
            ws = wb.active
            ws.title ='Hoja'+str(cont-1)
            bandera = False
        else:
            ws = wb.create_sheet('Hoja'+str(cont-1))
        
        
        ########
        for cand in candidatos:
            #controlador=4
            #ws.title ='Hoja'+str(cont)
            #ws = wb.create_sheet('Hoja'+str(cont))
            contador=1
            for letra in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                ws[letra+str(contador)].alignment = Alignment(horizontal = "center", vertical = "center")
                ws[letra+str(contador)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top  = Side(border_style = "thin"), bottom= Side(border_style = "thin"))
                ws[letra+str(contador)].fill = PatternFill(start_color ='66FFCC',end_color='66FFCC', fill_type='solid')
                ws[letra+str(contador)].font = Font(name ='Calibri', size=12, bold=True)
                ws.column_dimensions[letra].width = 20
        
            for letra in "ABCDEFGHIJKLM":
                ws["A"+letra+str(contador)].alignment = Alignment(horizontal = "center", vertical = "center")
                ws["A"+letra+str(contador)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top  = Side(border_style = "thin"), bottom= Side(border_style = "thin"))
                ws["A"+letra+str(contador)].fill = PatternFill(start_color ='66FFCC',end_color='66FFCC', fill_type='solid')
                ws["A"+letra+str(contador)].font = Font(name ='Calibri', size=12, bold=True)
                ws.column_dimensions["A"+letra].width = 20
            ws.row_dimensions[1].height = 30

            pais= str(cand.pais_nacimiento)
            if (cont==2):
                ws['A1'] = "NOMBRE"
                ws['B1'] = "ZONA"
                ws['C1'] = "Puesto"
                ws['D1'] = "Home Office"
                ws['E1'] = "Jornada"
                ws['F1'] = "Jornada Real"
                ws['G1'] = "Jornada en contrato"
                ws['H1'] = "Nacionalidad"
                ws['I1'] = "Edo civil"
                ws['J1'] = "Fecha de nacimiento"
                ws['K1'] = "CURP"
                ws['L1'] = "RFC"
                ws['M1'] = "Identificación expedida por"
                ws['N1'] = "Con número"
                ws['O1'] = "EMPRESA"
                ws['P1'] = "NSS"
                ws['Q1'] = "CALLE"
                ws['R1'] = "Número exterior"
                ws['S1'] = "Número interior"
                ws['T1'] = "COLONIA"
                ws['U1'] = "MUNICIPIO"
                ws['V1'] = "ESTADO"
                ws['W1'] = "CÓDIGO POSTAL"
                ws['X1'] = "NÚMERO DE TARJETA"
                ws['Y1'] = "NO DE CUENTA"
                ws['Z1'] = "CLABE INTERBANCARIA"
                ws['AA1'] = "NOMBRE DE BANCO"
                ws['AB1'] = "Tel. casa"
                ws['AC1'] = "Tel. personal"
                ws['AD1'] = "SD IMSS"
                ws['AE1'] = "Ayuda diaria"
                ws['AF1'] = "Sueldo neto"
                ws['AG1'] = "Cantidad letra"
                ws['AH1'] = "Bono"
                ws['AI1'] = "% comisión"
                ws['AJ1'] = "Fecha de ingreso"
                ws['AK1'] = "Fecha de alta IMSS"
                ws['AL1'] = "Correo empresa"
                ws['AM1'] = "Correo personal"
        
            ws['A'+str(cont)] = cand.nombre+" "+cand.apellido_paterno+" "+cand.apellido_materno
            ws['C'+str(cont)] = cand.puesto_solicitado
            ws['H'+str(cont)] = pais
            ws['I'+str(cont)] = cand.estado_civil
            ws['J'+str(cont)] = cand.fecha_nac
            ws['K'+str(cont)] = cand.curp
            ws['L'+str(cont)] = cand.rfc
            ws['P'+str(cont)] = cand.imss
            ws['Q'+str(cont)] = cand.calle
            ws['R'+str(cont)] = cand.num_ext
            ws['S'+str(cont)] = cand.num_int
            ws['T'+str(cont)] = cand.colonia
            ws['V'+str(cont)] = cand.esdo
            ws['W'+str(cont)] = cand.cp
            ws['AB'+str(cont)] = cand.tel
            ws['AM'+str(cont)] = cand.email_personal
            cont += 1
        

        #establecer nombre de archivo
        print("Salio")
        nombre_archivo = "ListaCandidatosExcel.xlsx"
        #definir tipo de respuesta
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


# Create your views here.  

def login_view(request):
    state = ""
    username = ""
    password = ""
    next = ""

    if request.GET:
        next = request.GET['next']

    if request.POST:
        username = request.POST['username']
        password = request.POST['password']

        if '@' in username:
            try:
                check = User.objects.get(email=username)
                username = check.username
            except:
                pass

        user = authenticate(username=username, password=password)

               

        if user is not None:
            if user.is_active:
                if user.groups.filter(name__in=['Administrador']).exists():
                    login(request, user)
                    return HttpResponseRedirect('/sadmin/')
                else:
                    if user.groups.filter(name__in=['RRHH']).exists():
                        login(request, user)
                        return HttpResponseRedirect('/rrhh/')
                    else:
                        if user.groups.filter(name__in=['Empleados']).exists():
                            login(request, user)
                            return HttpResponseRedirect('/cola-home/')
                        else:
                            return HttpResponseRedirect(next)

            else:#no activo
                state = 0
            #
            



            """if not user.is_superuser:
                if not user.groups.filter(name__in=['Becario']).exists():
                    if user.is_active:
                        login(request, user)
                        if next == "":
                            return HttpResponseRedirect('/portafolio/')
                        else:
                            return HttpResponseRedirect(next)
                    else:
                        state = 0
                else:
                    if user.is_active:
                      login(request, user)
                      if next == "":
                          return HttpResponseRedirect('/homebec/')



            else:
                if user.is_active:
                    login(request, user)
                    return HttpResponseRedirect('/sadmin/')
                else:
                    state = 0"""
        else:
            state = 0

    return render(request, 'dashboard/login.html', {'state':state, 'username': username, 'next': next,},)


####################################

@login_required(login_url = '/login/')
def crear(request, template_name = "dashboard/crear.html"):
    if request.method == 'POST':

        form = CreaEmpleado(request.POST)
        if form.is_valid():
             usuario = request.POST['username']
             first_name = request.POST['first_name']
             last_name = request.POST['last_name']
             correo = request.POST['email']
             contrasena = form.cleaned_data['password']
             usuario_inactivo = User.objects.get(email = correo)
             try:
                 if usuario_inactivo:
                     mensaje = 2
                 else:
                     # Crear nuevo USUARIO
                     user = User.objects.create_user(password=contrasena,is_superuser=False,username=usuario,first_name=first_name,last_name=last_name,email=correo,is_staff=False,is_active=False)
                     user.save(using=self._db)
                     mensaje = 1
                     return HttpResponseRedirect('/sadmin/')
             except:
                mensaje = 2
        else:
            print(form.errors)
    else:
        form = CreaEmpleado()
    return render(request, template_name, locals(),)


@login_required(login_url = '/login/')
def crear2(request, template_name = "dashboard/empAct.html"):
    if request.method == 'POST':

        form = CreaEmpleado2(request.POST)
        is_admin = request.POST['is_admin']
        username = request.POST['username']
        correo = request.POST['email']
        if  form.is_valid():
             try:
                    form.save()   
                    
                    if is_admin=="0":
                        group = Group.objects.get(name='Administrador')
                    elif is_admin=="1":
                        group = Group.objects.get(name='Consultor')
                    elif is_admin=="2":
                        group = Group.objects.get(name='Editor')
                    elif is_admin=="3":
                        group = Group.objects.get(name='Empleados')
                    else:
                        group = Group.objects.get(name='Becario')
                    us=User.objects.get(username = username, email= correo)
                    group.user_set.add(us)
                    form = CreaEmpleado2()
                    form.fields["username"].initial=""
                    form.fields["email"].initial=""
                    form.fields["first_name"].initial=""
                    form.fields["last_name"].initial=""
                    form.fields["password1"].initial=""
                    form.fields["password2"].initial=""
                    mensaje = 1
             except:
                 print("Error al guardar")
                 print(form.errors)
                 mensaje = 2
        else:
            print("Error de datos")
            #print(form)
            print(form.errors) 
    else:
        form = CreaEmpleado2()
    return render(request, template_name, locals(),) 

@login_required(login_url = '/login/')
def editarEmp(request, id, template_name = "dashboard/editarEmp.html"):
    if request.method == 'POST':

        form = CreaEmpleado2(request.POST)
        is_admin = request.POST['is_admin']
        username = request.POST['username']
        correo = request.POST['email']
        print(is_admin)
        if  form.is_valid():
             try:
                    form.save()    
                    if is_admin==0:
                        group = Group.objects.get(name='Administrador')
                    elif is_admin==1:
                        group = Group.objects.get(name='Consultor')
                    elif is_admin==2:
                        group = Group.objects.get(name='Editor')
                    elif is_admin==3:
                        group = Group.objects.get(name='Empleados')
                    else:
                        group = Group.objects.get(name='Becario')
                    us=User.objects.get(username = username, email= correo)
                    group.user_set.add(us)
                    mensaje = 1
             except:
                 print(form.errors)
                 mensaje = 2
        else:
            #print(form)
            print(form.errors)
    else:
        form = CreaEmpleado2()
    return render(request, template_name, locals(),)


@login_required(login_url = '/login/')
def sadmin(request, template_name = "dashboard/sadmin.html"):
    # Verificamos si el usuario tiene foto, Si existe algún registro en
    # Tabla Empleados que referencie al usuario en turno y si este posee
    # foto cargada al sistema
    usuario = request.user
    try:
        empleado = Empleado.objects.get(user = usuario.pk)
        foto = 1
        etapa = empleado.status
        if Estudio.objects.filter(user = usuario.pk).exists():
            profesion = Estudio.objects.get(user=usuario.pk)
        #empid=empleado
    except:
        foto = 0
        etapa = 1
        pass

    return render(request, template_name, locals(),)



@login_required(login_url = '/login/')
def rrhh(request, template_name = "dashboard/rrhh.html"):
    # Verificamos si el usuario tiene foto, Si existe algún registro en
    # Tabla Empleados que referencie al usuario en turno y si este posee
    # foto cargada al sistema
    usuario = request.user
    try:
        empleado = Empleado.objects.get(user = usuario.pk)
        foto = 1
        etapa = empleado.status
        if Estudio.objects.filter(user = usuario.pk).exists():
            profesion = Estudio.objects.get(user=usuario.pk)
        #empid=empleado
    except:
        foto = 0
        etapa = 1
        pass

    return render(request, template_name, locals(),)

def update_pro_docs(request):
    id=request.POST.get('est_Id')
   
    try:
        est=Estudios_pro.objects.get(pk=id)
        est.comprobante=est.comprobante=request.FILES['comprobante']
        est.save()
        response="OK"
    except Exception as e:
        response="ERROR"
        print(e)

    return HttpResponse(response)

def update_otro_docs(request):
    id=request.POST.get('est_Id')
    try:
        est=Estudios_otros.objects.get(pk=id)
        est.comprobante=request.FILES['comprobante']
        est.save()
        response="OK"
    except Exception as e:
        response="ERROR"
        print(e)

    return HttpResponse(response)

def update_docs(request):
    id=request.POST.get("cand_Id")
    curriculum=False
    inefrente=False
    ineatras=False
    acta=False
    curp=False
    rfc=False
    comprobante_domicilio=False
    imss=False
    carta1=False
    carta2=False
    contrato=False
    infonavit=False
    permiso=False
    oficio=False
    carta_menor=False
    INE_frenteTut=False
    INE_atrasTut=False
    
    #investigar si trae files
    if ("curriculum" in request.FILES):
        curriculum = request.FILES['curriculum']
    if ("INE_frente" in request.FILES):
        inefrente = request.FILES['INE_frente']
    if ("INE_atras" in request.FILES):
        ineatras = request.FILES['INE_atras']
    if ("acta" in request.FILES):
        acta = request.FILES['acta']
    if ("curp" in request.FILES):
        curp = request.FILES['curp']
    if ("rfc" in request.FILES):
        rfc = request.FILES['rfc']
    if ("comprobante_domicilio" in request.FILES):
        comprobante_domicilio = request.FILES['comprobante_domicilio']
    if ("imss" in request.FILES):
        imss = request.FILES['imss']
    if ("carta1" in request.FILES):
        carta1 = request.FILES['carta1']
    if ("carta2" in request.FILES):
        carta2 = request.FILES['carta2']
    if ("contrato" in request.FILES):
        contrato = request.FILES['contrato']
    if ("infonavit" in request.FILES):
        infonavit = request.FILES['infonavit']
    #extranjero
    if ("permiso" in request.FILES):
        permiso = request.FILES['permiso']
    #becario
    if ("oficio" in request.FILES):
        oficio = request.FILES['oficio']
    #menor
    if ("carta_menor" in request.FILES):
        carta_menor = request.FILES['carta_menor']
    if ("INE_frenteTut" in request.FILES):
        INE_frenteTut = request.FILES['INE_frenteTut']
    if ("INE_atrasTut" in request.FILES):
        INE_atrasTut = request.FILES['INE_atrasTut']

    
    cand = Candidato.objects.get(emp_id=id)
    if curriculum:
        cand.curriculum=curriculum
    if inefrente:
        cand.docu_ident_front=inefrente
    if ineatras:
        cand.docu_ident_back=ineatras
    if acta:
        cand.acta_nacimiento=acta
    if curp:
        cand.imagen_curp=curp
    if rfc:
        cand.imagen_rfc=rfc
    if comprobante_domicilio:
        cand.comprobante_domicilio=comprobante_domicilio
    if imss:
        cand.imagen_imss=imss
    if carta1:
        cand.carta1_recomendacion=carta1
    if carta2:
        cand.carta2_recomendacion=carta2
    if contrato:
        cand.contrato=contrato
    if infonavit:
        cand.imagen_infonavit=infonavit
    if permiso:
        cand.permiso=permiso
    if oficio:
        cand.oficio=oficio
    if carta_menor:
        cand.carta_menor=carta_menor
    if INE_frenteTut:
        cand.INE_frenteTut=INE_frenteTut
    if INE_atrasTut:
        cand.INE_atrasTut=INE_atrasTut
    
    print(inefrente)
    try:
        cand.save()
        response="OK"
    except Exception as e:
        response="ERROR"
        print(e)
    return HttpResponse(response)
        


@login_required(login_url = '/login/')
def candDocs(request,id, template_name = "dashboard/candDocs.html"):
    candId=id
    curriculum=False
    inefrente=False
    ineatras=False
    acta=False
    curp=False
    rfc=False
    comprobante=False
    imss=False
    carta1=False
    carta2=False
    contrato=False
    infonavit=False
    permiso=False
    oficio=False
    carta_menor=False
    INE_frenteTut=False
    INE_atrasTut=False
    #lifrente=False
    #licatras=False
    if (request.POST):
        #si hay post
        licatras=False

    
    else:
        user = request.user
        empleado = Empleado.objects.filter(user = user.pk)
        if (empleado):
            foto = 1
        else:
            foto = 0
    
        candidato = Candidato.objects.get(emp_id=id)
        #cand= Candidato.objects.get(emp_id=id)
        
        if (candidato.curriculum not in [None,'']):
            curriculum=candidato.curriculum
        if (candidato.docu_ident_front not in [None,'']):
            inefrente=candidato.docu_ident_front
        if (candidato.docu_ident_back not in [None,'']):
            ineatras=candidato.docu_ident_back
        if (candidato.acta_nacimiento not in [None,'']):
            acta=candidato.acta_nacimiento
        if (candidato.imagen_curp not in [None,'']):
            curp=candidato.imagen_curp
        if (candidato.imagen_rfc not in [None,'']):
            rfc=candidato.imagen_rfc
        if (candidato.comprobante_domicilio not in [None,'']):
            comprobante=candidato.comprobante_domicilio
        if (candidato.imagen_imss not in [None,'']):
            imss=candidato.imagen_imss
        if (candidato.carta1_recomendacion not in [None,'']):
            carta1=candidato.carta1_recomendacion
        if (candidato.carta2_recomendacion not in [None,'']):
            carta2=candidato.carta2_recomendacion
        if (candidato.contrato not in [None,'']):
            contrato=candidato.contrato
        if (candidato.imagen_infonavit not in [None,'']):
            infonavit=candidato.imagen_infonavit
        if (candidato.permiso not in [None,'']):
            permiso=candidato.permiso
        if (candidato.oficio not in [None,'']):
            oficio=candidato.oficio
        if (candidato.carta_menor not in [None,'']):
            carta_menor=candidato.carta_menor
        if (candidato.INE_frenteTut not in [None,'']):
            INE_frenteTut=candidato.INE_frenteTut
        if (candidato.INE_frenteTut not in [None,'']):
            INE_atrasTut=candidato.INE_atrasTut
        
        if int(candidato.edad)<18:
            menor=True
        else:
            menor=False
        
        if str(candidato.pais_nacimiento)=="MÉXICO":
            extranjero=False
        else:
            extranjero=True
       
        if User.objects.filter(pk=id, groups__name='Becario').exists():
            becario=True
        else:
            becario=False
        

        form_est=Formulario_est_comp()
        form_est.fields['comprobante'].widget.attrs['accept'] ='image/*'
        form_est.fields['comprobante'].widget.attrs['class'] ='comprobante'
    
        form=Formulario_cand_docs()
        form.fields['curriculum'].widget.attrs['accept'] ='application/pdf'
        form.fields['INE_frente'].widget.attrs['accept'] ='image/*'
        form.fields['INE_atras'].widget.attrs['accept'] ='image/*'
        form.fields['acta'].widget.attrs['accept'] ='image/*'
        form.fields['curp'].widget.attrs['accept'] ='image/*'
        form.fields['rfc'].widget.attrs['accept'] ='image/*'
        form.fields['comprobante_domicilio'].widget.attrs['accept'] ='image/*'
        form.fields['imss'].widget.attrs['accept'] ='image/*'
        form.fields['carta1'].widget.attrs['accept'] ='image/*'
        form.fields['carta2'].widget.attrs['accept'] ='image/*'
        form.fields['contrato'].widget.attrs['accept'] ='image/*'
        form.fields['infonavit'].widget.attrs['accept'] ='image/*'
        #extranjero
        form.fields['permiso'].widget.attrs['accept'] ='image/*'
        #becario
        form.fields['oficio'].widget.attrs['accept'] ='image/*'
        #menor de edad
        form.fields['carta_menor'].widget.attrs['accept'] ='image/*'
        form.fields['INE_frenteTut'].widget.attrs['accept'] ='image/*'
        form.fields['INE_atrasTut'].widget.attrs['accept'] ='image/*'
        #estudios profesionales
        Est_pro=Estudios_pro.objects.filter(candidato=candidato)
        #otros estudios
        Est_otros=Estudios_otros.objects.filter(candidato=candidato)

   
    return render(request, template_name, locals(),)

@login_required(login_url = '/login/')
def candLst(request, template_name = "dashboard/candLst.html"):
    
    user = request.user
    empleado = Empleado.objects.filter(user = user.pk)
    if (empleado):
        foto = 1
    else:
        foto = 0
    
    candidatos = Candidato.objects.filter(emp_id=None)
    
    ###form para Docs
    
    #form=Formulario_cand_docs()
    
    #idiomas = Idioma.objects.filter(id=candidato)
    #estudios = Estudios_pro.objects.filter(id=candidato)
    #hermanos = Hermano_candidato.objects.filter(id=candidato)
    #hijos = Hijo_candidato.objects.filter(id=candidato)
    #hijos = Hijo_candidato.objects.filter(id=candidato)
    #experiencias = Experiencia.objects.filter(id=candidato)
    #referencias = Referencia.objects.filter(id=candidato)
    #empleados = User.objects.filter(is_superuser=0,is_active=1)
    return render(request, template_name, locals(),)




@login_required(login_url = '/login/')
def candrrhh(request,id, template_name = "candidatos/captura_cv.html"):
    rrhh="RRHH"
    status_con="True"
    con_id=id
    #llenar candidato
    candidato= Candidato.objects.get(id=id)
    #form_candidato=Formulario_candidato()
    form_candidato=set_values_candidato(candidato)
    form_candidato.fields['con_id'].initial = con_id

    form_idioma=Formulario_idioma()
    form_editidioma=Formulario_idioma()
    form_hermano=Formulario_hermano_candidato()
    form_edithermano=Formulario_hermano_candidato()
    form_hijo=Formulario_hijo_candidato
    form_edithijo=Formulario_hijo_candidato
    form_referencia=Formulario_referencia()
    form_editreferencia=Formulario_referencia()
    form_experiencia=Formulario_experiencia()
    form_editexperiencia=Formulario_experiencia()
    form_estudio=Formulario_estudios()
    form_editestudio=Formulario_estudios()
    form_estudiootro=Formulario_estudiosotros()
    form_editestudiootro=Formulario_estudiosotros()
        
    
    experiencias = Experiencia.objects.filter(candidato=candidato)
    referencias = Referencia.objects.filter(candidato=candidato)
    hijos = Hijo_candidato.objects.filter(candidato=candidato)
    hermanos = Hermano_candidato.objects.filter(candidato=candidato)
    idiomas = Idioma_candidato.objects.filter(candidato=candidato)
    estudios = Estudios_pro.objects.filter(candidato=candidato)
        
    form_idioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"
    form_editidioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"

        

    form_candidato.fields['edad'].widget.attrs['readonly'] ='readonly'
    form_candidato.fields['piso'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['depto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_que'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_donde'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_horario'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_termino'].widget.attrs['disabled'] ='disabled'

        
        
    form_candidato.fields['primaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['secundaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['preparatoria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_documento'].widget.attrs['disabled'] ='disabled'
      
    form_candidato.fields['tecnica_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_documento'].widget.attrs['disabled'] ='disabled'

    

    form_candidato.fields['pago_infonavit'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_marca'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_modelo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['seguro_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['afianzado_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_cargo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_fuente'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_depto'].widget.attrs['disabled'] ='disabled'

    form_experiencia.fields['experiencia_supervision_num'].widget.attrs['disabled'] ='disabled'
    

    return render(request, template_name, locals(),)

def buscar_cand(request):
    data = dict()
    id=request.GET.get('id')
    data['form_is_valid'] = False
    try:
        cand = Candidato.objects.get(emp_id = id)
        data['form_is_valid'] = True
    except Exception as e:
        pass
        
    

    return JsonResponse(data)

@login_required(login_url = '/login/')
def empAct(request, template_name = "dashboard/empAct.html"):
    user = request.user
    
    empleado = Empleado.objects.filter(user = user.pk)
    if (empleado):
        foto = 1
    else:
        foto = 0
    if request.method == 'POST':
         try:
            tipoemp = TipoEmpleado(request.POST)
            if tipoemp.is_valid():
                tipo = request.POST["tipo_empleado"]

                if tipo=="1":
                    grupo = Group.objects.get(name="Empleados")
                elif tipo=="2":
                    grupo = Group.objects.get(name="Editor")
                elif tipo=="3":
                    grupo = Group.objects.get(name="Consultor")
                elif tipo=="4":
                    grupo = Group.objects.get(name="Becario")
                else:
                    grupo = Group.objects.get(name="Administrador")

                empleados =grupo.user_set.all()
                form = TipoEmpleado(request.POST)# TipoEmpleado()
                form_emp = EditEmpleado()
                nform = CreaEmpleado2()

         except Exception as e:
             print("error existe")
             print(e) 
             pass
             
    else:
        
        grupo = Group.objects.get(name="Empleados")
        empleados = grupo.user_set.all()
        form = TipoEmpleado()
        form_emp = EditEmpleado()
        nform = CreaEmpleado2()

    #empleados = User.objects.filter(is_superuser=0,is_active=1)
    

    return render(request, template_name, locals(),)

@login_required(login_url = '/login/')
def empLst(request, template_name = "dashboard/empLst.html"):
    user = request.user
    
    empleado = Empleado.objects.filter(user = user.pk)
    if (empleado):
        foto = 1
    else:
        foto = 0
    if request.method == 'POST':
         try:
            tipoemp = TipoEmpleado(request.POST)
            if tipoemp.is_valid():
                tipo = request.POST["tipo_empleado"]

                if tipo=="1":
                    grupo = Group.objects.get(name="Empleados")
                elif tipo=="2":
                    grupo = Group.objects.get(name="Editor")
                elif tipo=="3":
                    grupo = Group.objects.get(name="Consultor")
                elif tipo=="4":
                    grupo = Group.objects.get(name="Becario")
                elif tipo=="5":
                    grupo = Group.objects.get(name="Administrador")
                else:
                    grupo = Group.objects.get(name="RRHH")

                empleados =grupo.user_set.filter(is_active=true)
                form = TipoEmpleado(request.POST)# TipoEmpleado()
                form_emp = EditEmpleado()
                nform = CreaEmpleado2()
                ###form para Docs
                form_doc=Formulario_cand_docs()

         except Exception as e:
             print(e) 
             pass
             
    else:
        
        grupo = Group.objects.get(name="Empleados")
        empleados = grupo.user_set.filter(is_active=true)
        form = TipoEmpleado()
        form_emp = EditEmpleado()
        nform = CreaEmpleado2()
        ###form para Docs
        form_doc=Formulario_cand_docs()
        ###form para posible baja
        form_baja=Formulario_baja()


    #empleados = User.objects.filter(is_superuser=0,is_active=1)
    

    return render(request, template_name, locals(),)


@login_required(login_url = '/login/')
def empIna(request, template_name = "dashboard/empIna.html"):
    grupo = Group.objects.get(name="Empleados")
    empleados =grupo.user_set.all().filter(is_superuser=0,is_active=0)
    #empleados = User.objects.filter(is_superuser=3 , is_active=0)
    

    return render(request, template_name, locals(),)

@login_required(login_url = '/login/')
def admAct(request, template_name = "dashboard/admAct.html"):
    grupo = Group.objects.get(name="Administrador")
    empleados =grupo.user_set.all().filter(is_superuser=1,is_active=1)
    #empleados = User.objects.filter(is_superuser=1)
    

    return render(request, template_name, locals(),)

@login_required(login_url = '/login/')
def admBec(request, template_name = "dashboard/admBec.html"):
    grupo = Group.objects.get(name="Becario")
    empleados =grupo.user_set.all().filter(is_superuser=0)
    #empleados = User.objects.filter(is_superuser=4)
    

    return render(request, template_name, locals(),)

######################################## 
    ############## Visor de etapas del empleado
@login_required(login_url = '/login/')
def dashAdm(request,id, template_name = "dashboard/dashboardAdm.html"):
    #usuario = request.user
    try:
        #print(id)
        empleado = Empleado.objects.get(user_id = id)
        usuario = User.objects.get(id = empleado.pk)
        etapa = empleado.status
        print(etapa)
        print(empleado.user_id)
        #empleado.status = 2 
        #empleado.save()
    except:
        pass
    return render(request, template_name, locals(),)




    ############### 


 
@login_required(login_url = '/login/')
def login_out(request):
    logout(request)
    return HttpResponseRedirect('/login/')

@login_required(login_url = '/login/')
def porta_emp(request, template_name = "candidatos/captura_cv.html"):
    # Verificamos si el usuario tiene foto, Si existe algún registro en
    # Tabla Empleados que referencie al usuario en turno y si este posee
    # foto cargada al sistema
    usuario = request.user
    try:
        empleado = Empleado.objects.get(user = usuario.pk)
        foto = 1
        etapa = empleado.status
        if Estudio.objects.filter(user = usuario.pk).exists():
            profesion = Estudio.objects.get(user=usuario.pk)
    except:
        foto = 0
        etapa = 1
        pass
    #investigar si tiene archivo en candidatos
    #llenar candidato
    rrhh="PORTA_EMP"
    emp_id=usuario.id
    form_candidato=Formulario_candidato()
    try:
        candidato= Candidato.objects.get(emp_id=usuario.id)
        status_con="True"
        con_id=candidato.id
        form_candidato.fields['con_id'].initial = con_id
        #form_candidato=Formulario_candidato()
        form_candidato=set_values_candidato(candidato)
        experiencias = Experiencia.objects.filter(candidato=candidato)
        referencias = Referencia.objects.filter(candidato=candidato)
        hijos = Hijo_candidato.objects.filter(candidato=candidato)
        hermanos = Hermano_candidato.objects.filter(candidato=candidato)
        idiomas = Idioma_candidato.objects.filter(candidato=candidato)
        estudios = Estudios_pro.objects.filter(candidato=candidato)
        estudiosotros=Estudios_otros.objects.filter(candidato=candidato)
    except Exception as e:
        pass
    ####FORMULARIOS
    form_idioma=Formulario_idioma()
    form_editidioma=Formulario_idioma()
    form_hermano=Formulario_hermano_candidato()
    form_edithermano=Formulario_hermano_candidato()
    form_hijo=Formulario_hijo_candidato
    form_edithijo=Formulario_hijo_candidato
    form_referencia=Formulario_referencia()
    form_editreferencia=Formulario_referencia()
    form_experiencia=Formulario_experiencia()
    form_editexperiencia=Formulario_experiencia()
    form_estudio=Formulario_estudios()
    form_editestudio=Formulario_estudios()
    form_estudiootro=Formulario_estudiosotros()
    form_editestudiootro=Formulario_estudiosotros()

    form_idioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"
    form_editidioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"

        

    form_candidato.fields['edad'].widget.attrs['readonly'] ='readonly'
    form_candidato.fields['piso'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['depto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_que'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_donde'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_horario'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_termino'].widget.attrs['disabled'] ='disabled'

        
        
    form_candidato.fields['primaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['secundaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['preparatoria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_documento'].widget.attrs['disabled'] ='disabled'
      
    form_candidato.fields['tecnica_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_documento'].widget.attrs['disabled'] ='disabled'

    

    form_candidato.fields['pago_infonavit'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_marca'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_modelo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['seguro_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['afianzado_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_cargo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_fuente'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_depto'].widget.attrs['disabled'] ='disabled'

    form_experiencia.fields['experiencia_supervision_num'].widget.attrs['disabled'] ='disabled'
    

    
    
    return render(request, template_name, locals(),)

@login_required(login_url = '/login/')
def portafolio(request, template_name = "candidatos/captura_cv.html"):
    # Verificamos si el usuario tiene foto, Si existe algún registro en
    # Tabla Empleados que referencie al usuario en turno y si este posee
    # foto cargada al sistema
    usuario = request.user
    try:
        empleado = Empleado.objects.get(user = usuario.pk)
        foto = 1
        etapa = empleado.status
        if Estudio.objects.filter(user = usuario.pk).exists():
            profesion = Estudio.objects.get(user=usuario.pk)
    except:
        foto = 0
        etapa = 1
        pass
    #investigar si tiene archivo en candidatos
    #llenar candidato
    rrhh="PORTAFOLIO"
    emp_id=usuario.id
    form_candidato=Formulario_candidato()
    try:
        candidato= Candidato.objects.get(emp_id=usuario.id)
        status_con="True"
        con_id=candidato.id
        form_candidato.fields['con_id'].initial = con_id
        #form_candidato=Formulario_candidato()
        form_candidato=set_values_candidato(candidato)
        experiencias = Experiencia.objects.filter(candidato=candidato)
        referencias = Referencia.objects.filter(candidato=candidato)
        hijos = Hijo_candidato.objects.filter(candidato=candidato)
        hermanos = Hermano_candidato.objects.filter(candidato=candidato)
        idiomas = Idioma_candidato.objects.filter(candidato=candidato)
        estudios = Estudios_pro.objects.filter(candidato=candidato)
        estudiosotros=Estudios_otros.objects.filter(candidato=candidato)
    except Exception as e:
        pass
    ####FORMULARIOS
    form_idioma=Formulario_idioma()
    form_editidioma=Formulario_idioma()
    form_hermano=Formulario_hermano_candidato()
    form_edithermano=Formulario_hermano_candidato()
    form_hijo=Formulario_hijo_candidato
    form_edithijo=Formulario_hijo_candidato
    form_referencia=Formulario_referencia()
    form_editreferencia=Formulario_referencia()
    form_experiencia=Formulario_experiencia()
    form_editexperiencia=Formulario_experiencia()
    form_estudio=Formulario_estudios()
    form_editestudio=Formulario_estudios()
    form_estudiootro=Formulario_estudiosotros()
    form_editestudiootro=Formulario_estudiosotros()

    form_idioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"
    form_editidioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"

        

    form_candidato.fields['edad'].widget.attrs['readonly'] ='readonly'
    form_candidato.fields['piso'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['depto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_que'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_donde'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_horario'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_termino'].widget.attrs['disabled'] ='disabled'

        
        
    form_candidato.fields['primaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['secundaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['preparatoria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_documento'].widget.attrs['disabled'] ='disabled'
      
    form_candidato.fields['tecnica_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_documento'].widget.attrs['disabled'] ='disabled'

    

    form_candidato.fields['pago_infonavit'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_marca'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_modelo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['seguro_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['afianzado_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_cargo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_fuente'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_depto'].widget.attrs['disabled'] ='disabled'

    form_experiencia.fields['experiencia_supervision_num'].widget.attrs['disabled'] ='disabled'
    

    
    
    return render(request, template_name, locals(),)




def register(request, template_name = "dashboard/register.html"):
    if request.method == 'POST':
        form = ActivaEmpleado(request.POST)
        if form.is_valid():
            correo = form.cleaned_data['email']
            usuario = form.cleaned_data['username']
            contrasena = form.cleaned_data['password']
            # Consultar si existe correo
            try:
                usuario_inactivo = User.objects.get(email = correo)
                if usuario_inactivo:
                    # Verificar su status (ACTIVADO O NO ACTIVADO)
                    if usuario_inactivo.is_active:
                        mensaje = 0
                        form = ActivaEmpleado()
                        form.fields["username"].initial=""
                        form.fields["email"].initial=""
                    else:
                        # Actualizar USERNAME
                        usuario_inactivo.username = usuario
                        usuario_inactivo.save()
                        # Actualizar PASSWORD
                        usuario_inactivo.set_password(contrasena)
                        usuario_inactivo.save()
                        # Cambiar status
                        usuario_inactivo.is_active = True
                        usuario_inactivo.save()
                        mensaje = 1
                        form = ActivaEmpleado()
                        form.fields["username"].initial=""
                        form.fields["email"].initial=""
                        
                        
            except:
                form = ActivaEmpleado()
                form.fields["username"].initial=""
                form.fields["email"].initial=""
                mensaje = 2
                
                
        else:
            form = ActivaEmpleado()
            form.fields["username"].initial=""
            form.fields["email"].initial=""
            print(form.errors)
    else:
        form = ActivaEmpleado()
    return render(request, template_name, locals(),)

def update_per(request):
    response ="NONE"
    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        emp_id = request.POST.get('id')
        if (password1==password2):
            username = request.POST.get('username')
            user_existente = User.objects.get(username = username)
            if (user_existente):
                if(user_existente.id != int(emp_id)):
                    response="user_existe"
                    return HttpResponse(response)
            print(request.POST)
            email = request.POST['email']
            print(email)
            email_existente = User.objects.get(email = email)
            if (email_existente):
                if(email_existente.id != int(emp_id)):
                    response="email_existe"
                    return HttpResponse(response)
            try:
                username = request.POST.get('username')
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                email = request.POST.get('email')
                is_active = request.POST.get('is_active')
                is_superuser = False
                password = request.POST.get('password1')
                per= User.objects.get(id=emp_id)
                per.username=username
                per.first_name=first_name
                per.last_name=last_name
                per.email=email
                per.is_active=is_active
                per.is_superuser=is_superuser
                per.set_password(password)
                per.save()
                response ="OK"
            except:
                print(per.errors)
                response="error"
        else:
            response="passerror"
        
    return HttpResponse(response)

def new_per(request):
    response="NONE"
    if request.method == 'POST':
        form = CreaEmpleado2(request.POST)
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        if (password1==password2):
            
            username = request.POST.get('username')
            user_existente = User.objects.filter(username = username).exists()
            
            if not user_existente:
                email = request.POST['email']
                email_existente = User.objects.filter(email = email).exists()
                if not email_existente:
                    if form.is_valid():
                        try:
                            form.save()
                            is_admin = request.POST['is_admin']
                            if is_admin=="0":
                                group = Group.objects.get(name='Administrador')
                            elif is_admin=="1":
                                group = Group.objects.get(name='RRHH')
                            elif is_admin=="2":
                                group = Group.objects.get(name='Editor')
                            elif is_admin=="3":
                                group = Group.objects.get(name='Empleados')
                            elif is_admin=="4":
                                group = Group.objects.get(name='Becario')
                            else:
                                group = Group.objects.get(name='Consultor')
                            us=User.objects.get(username = username, email= email)
                            group.user_set.add(us)
                            data = {
                                'result': 'OK'
                            }
                            response=data
                        except:
                            print(form.errors)
                            data = {
                                'result': 'error'
                            }
                            response=data
                    else:
                        print(form.errors)
                        #pas_err=form.errors
                        
                        if form.errors['password2']:
                            pas_err=form.errors['password2']
                            data = {
                                'result': 'password2',
                                'message': pas_err
                            }
                            
                            #pas_err=form.errors.as_text
                        
                        response=data
                else:
                    print(form.errors)
                    data = {
                        'result': 'email_existe',
                    }
                    response=data
            else:
                print(form.errors)
                data = {
                     'result': 'user_existe'
                }
                response=data
        else:
            print(form.errors)
            data = {
                'result': 'passerror'
            }
            response=data
            
        
       
    #return HttpResponse(response)
    return JsonResponse(response)



def add_per2(request):
    if request.method == 'POST':
        print("entro a post")

        form = CreaEmpleado2(request.POST)
        is_admin = request.POST['is_admin']
        username = request.POST['username']
        correo = request.POST['email']
        response=""
        if  form.is_valid():
             try:
                    #form.save()   
                    
                    if is_admin=="0":
                        group = Group.objects.get(name='Administrador')
                    elif is_admin=="1":
                        group = Group.objects.get(name='Consultor')
                    elif is_admin=="2":
                        group = Group.objects.get(name='Editor')
                    elif is_admin=="3":
                        group = Group.objects.get(name='Empleados')
                    else:
                        group = Group.objects.get(name='Becario')
                    #us=User.objects.get(username = username, email= correo)
                    #group.user_set.add(us)
                    
                    #form = CreaEmpleado2()
                    #form.fields["username"].initial=""
                    #form.fields["email"].initial=""
                    #form.fields["first_name"].initial=""
                    #form.fields["last_name"].initial=""
                    #form.fields["password1"].initial=""
                    #form.fields["password2"].initial=""
                    response="OK"
             except:
                 response="error"
                 pass
               
        else:
            response="existe"
        return HttpResponse(response)
            #print(form)
            #print("error en valid")
            #print(form.errors) 
   
      # manda existe en add_per

    #return HttpResponseRedirect(self.get_success_url())
    #display_html_dict = self._displayhtml(request)
    #return render(request,self.display_html_dict)
    #return JsonResponse({
    #            'success':False,
    #            'err_code':'invalid_form',
    #            'err_msg':form.errors
    #        })


def filtra_cand(request):
    data = dict()
    palabra=request.GET.get('dato')
    #name__iexact="beatles blog" __icontains
    candidatos = Candidato.objects.filter(puesto_solicitado__icontains=palabra,emp_id=None)
    if candidatos:
        data['form_is_valid'] = True  #si encontro registros
        data['html_candidatos_lista'] = render_to_string('dashboard/cand_cont.html', {
        'candidatos': candidatos
        })
    else:
        data['form_is_valid'] = False  #si no encontro registros
    
    return JsonResponse(data)

def filtra2_cand(request):
    data = dict()
    nom=request.GET.get('nom')
    pat=request.GET.get('pat')
    mat=request.GET.get('mat')
    #name__iexact="beatles blog" __icontains
    candidatos = Candidato.objects.filter(nombre__icontains=nom,apellido_paterno__icontains=pat,apellido_materno__icontains=mat,emp_id=None)
    if candidatos:
        data['form_is_valid'] = True  #si encontro registros
        data['html_candidatos_lista'] = render_to_string('dashboard/cand_cont.html', {
        'candidatos': candidatos
        })
    else:
        data['form_is_valid'] = False  #si no encontro registros
    
    return JsonResponse(data)

def filtra3_cand(request):
    data = dict()
    fec=request.GET.get('fec')
    fec=mod_fecha(fec)
    print(fec)
    candidatos = Candidato.objects.filter(fecha_solicitud__date=fec,emp_id=None)
    if candidatos:
        data['form_is_valid'] = True  #si encontro registros
        data['html_candidatos_lista'] = render_to_string('dashboard/cand_cont.html', {
        'candidatos': candidatos
        })
    else:
        data['form_is_valid'] = False  #si no encontro registros
    
    return JsonResponse(data)

def all_cand(request):
    data = dict()
    candidatos = Candidato.objects.filter(emp_id=None)
    if candidatos:
        data['form_is_valid'] = True  #si encontro registros
        data['html_candidatos_lista'] = render_to_string('dashboard/cand_cont.html', {
        'candidatos': candidatos
        })
    else:
        data['form_is_valid'] = False  #si no encontro registros
    
    return JsonResponse(data)

def del_candidato(request):#elimina candidato
    data = dict()
    id=request.GET.get('id')
    
    candidato= Candidato.objects.get(pk=id)
    candidato.delete()
    data['form_is_valid'] = True
    candidatos = Candidato.objects.filter(emp_id=None)
    data['html_candidatos_lista'] = render_to_string('dashboard/cand_cont.html', {
        'candidatos': candidatos
    })
    return JsonResponse(data)

@login_required(login_url = '/login/')
def pasar_emp(request,id, template_name = "dashboard/pasar_emp.html"):
    idcand=id
    nform = set_values_emp(idcand)

    return render(request, template_name, locals(),)

def new_emp(request):
    response="NONE"
    if request.method == 'POST':
        form = CreaEmpleado2(request.POST)
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        idcand = request.POST.get('idcand')
        
        if (password1==password2):
            
            username = request.POST.get('username')
            user_existente = User.objects.filter(username = username).exists()
            
            if not user_existente:
                email = request.POST['email']
                email_existente = User.objects.filter(email = email).exists()
                if not email_existente:
                    if form.is_valid():
                        try:
                            #form.save()
                            post=form.save()
                            group = Group.objects.get(name='Empleados')
                            us=User.objects.get(username = username, email= email)
                            group.user_set.add(us)
                            us.is_active=True
                            us.save()
                            data = {
                                'result': 'OK'
                            }
                            response=data
                            #actualizar modelo de candidato emp_id
                            cand=Candidato.objects.get(id=idcand)
                            cand.emp_id=post.id
                            cand.save()
                            ################### enviar correo
                            message="Se ha generado una cuenta de acceso a su nombre usuario : "+username+" password: "+password1
                            name="Recursos humanos"
                            email='rrhh@springlabs.net'
                            ##correo de candidato personal
                            destino=cand.email_personal

                            print(destino)
                            
                            body=render_to_string(
                                'email_content.html',{
                                'name':name,
                                'email':email,
                                'message':message,
                                },
                            )
                            email_message = EmailMessage(
                                subject='Mensaje de usuario',
                                body=body,
                                from_email=email,
                                #to=['rrhh@springlabs.net'],
                                to=[destino],
                            )
                            email_message.content_subtype = 'html'
                            email_message.send()
                        except Exception as e:
                            print(e)
                            print(form.errors)
                            data = {
                                'result': 'error'
                            }
                            response=data
                    else:
                        #print(form.errors)
                        for er in form.errors:
                            print(er)
                            pas_err=form.errors[er]
                            data = {
                                'result': er,
                                'message': pas_err
                            }
                            response=data
                else:
                    print(form.errors)
                    data = {
                        'result': 'email_existe',
                    }
                    response=data
            else:
                print(form.errors)
                data = {
                     'result': 'user_existe'
                }
                response=data
        else:
            print(form.errors)
            data = {
                'result': 'passerror'
            }
            response=data
            
        
       
    #return HttpResponse(response)
    return JsonResponse(response)

def baja_emp(request):#baja de empleado
    data = dict()
    id=request.GET.get('emp_Id')
    motivo=request.GET.get('motivo')
    fecha = datetime.strptime(request.GET.get('fecha_baja'), '%d/%m/%Y').strftime('%Y-%m-%d')
    tipo=request.GET.get('tipo_baja')
    try:
        #modificar en tabla candidato activo x NO
        emp= Candidato.objects.get(emp_id=id)
        emp.motivo=motivo
        emp.fecha_baja=fecha
        emp.tipo_baja=tipo
        emp.activo='NO'
        emp.save()
        #modificar en tabla user is_active a false
        usr= User.objects.get(pk=emp.emp_id)
        grupo = Group.objects.get(name="Ex-Empleados")
        grupo.user_set.add(usr)
        grupo2 = Group.objects.get(name="Empleados")
        grupo2.user_set.remove(usr)
        usr.is_active=False
        usr.save()

        data['form_is_valid'] = True
        grupo = Group.objects.get(name="Empleados")
        empleados = grupo.user_set.filter(is_active=true)
        #empleados = Candidato.objects.filter(activo="SI")
        data['html_empleados_lista'] = render_to_string('dashboard/emp_cont.html', {
            'empleados': empleados
        })
    except Exception as e:
        data['form_is_valid'] = False
        print(e)
    
    

    
    return JsonResponse(data)

def home(request, template_name = "dashboard/home.html"):
# Imagen de <a href="https://pixabay.com/es/users/geralt-9301/?utm_source=link-attribution&amp;utm_medium=referral&amp;utm_campaign=image&amp;utm_content=2499628">Gerd Altmann</a> en <a href="https://pixabay.com/es/?utm_source=link-attribution&amp;utm_medium=referral&amp;utm_campaign=image&amp;utm_content=2499628">Pixabay</a>
    return render(request, template_name, locals(),)


@login_required(login_url = '/login/')
def cola_home(request, template_name = "dashboard/cola_home.html"):
    # Verificamos si el usuario tiene foto, Si existe algún registro en
    # Tabla Empleados que referencie al usuario en turno y si este posee
    # foto cargada al sistema
    usuario = request.user
    try:
        empleado = Empleado.objects.get(user = usuario.pk)
        foto = 1
    #    etapa = empleado.status
        if Estudio.objects.filter(user = usuario.pk).exists():
            profesion = Estudio.objects.get(user=usuario.pk)
        #empid=empleado
    except:
        foto = 0
        etapa = 1
        pass

    return render(request, template_name, locals(),)


#### COLABORADOR#########################################################
#### COLABORADOR#########################################################
#### COLABORADOR#########################################################
@login_required(login_url = '/login/')
def colaborador(request,id, template_name = "colaboradores/colaborador.html"):
    rrhh="COLABORADOR"
    status_con="True"
    #llenar candidato
    candidato= Candidato.objects.get(emp_id=id)
    con_id=candidato.id
    #form_candidato=Formulario_candidato()
    form_candidato=set_values_candidato(candidato)

    
    form_candidato.fields['con_id'].initial = con_id

    form_idioma=Formulario_idioma()
    form_editidioma=Formulario_idioma()
    form_hermano=Formulario_hermano_candidato()
    form_edithermano=Formulario_hermano_candidato()
    form_hijo=Formulario_hijo_candidato
    form_edithijo=Formulario_hijo_candidato
    form_referencia=Formulario_referencia()
    form_editreferencia=Formulario_referencia()
    form_experiencia=Formulario_experiencia()
    form_editexperiencia=Formulario_experiencia()
    form_estudio=Formulario_estudios()
    form_editestudio=Formulario_estudios()
    
    form_estudiootro=Formulario_estudiosotros()
    form_editestudiootro=Formulario_estudiosotros()
        
    
    experiencias = Experiencia.objects.filter(candidato=candidato)
    referencias = Referencia.objects.filter(candidato=candidato)
    hijos = Hijo_candidato.objects.filter(candidato=candidato)
    hermanos = Hermano_candidato.objects.filter(candidato=candidato)
    idiomas = Idioma_candidato.objects.filter(candidato=candidato)
    estudios = Estudios_pro.objects.filter(candidato=candidato)
    estudiosotros=Estudios_otros.objects.filter(candidato=candidato)
        
    form_idioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"
    form_editidioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"

        

    form_candidato.fields['edad'].widget.attrs['readonly'] ='readonly'
    form_candidato.fields['piso'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['depto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_que'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_donde'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_horario'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_termino'].widget.attrs['disabled'] ='disabled'

        
        
    form_candidato.fields['primaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['secundaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['preparatoria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_documento'].widget.attrs['disabled'] ='disabled'
      
    form_candidato.fields['tecnica_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_documento'].widget.attrs['disabled'] ='disabled'

    

    form_candidato.fields['pago_infonavit'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_marca'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_modelo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['seguro_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['afianzado_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_cargo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_fuente'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_depto'].widget.attrs['disabled'] ='disabled'

    form_experiencia.fields['experiencia_supervision_num'].widget.attrs['disabled'] ='disabled'
    

    return render(request, template_name, locals(),)

@login_required(login_url = '/login/')
def colarrhh(request,id, template_name = "candidatos/captura_cv.html"):
    rrhh="COLABORADOR"
    status_con="True"
    #llenar candidato
    candidato= Candidato.objects.get(emp_id=id)
    con_id=candidato.id
    #form_candidato=Formulario_candidato()
    form_candidato=set_values_candidato(candidato)

    
    form_candidato.fields['con_id'].initial = con_id

    form_idioma=Formulario_idioma()
    form_editidioma=Formulario_idioma()
    form_hermano=Formulario_hermano_candidato()
    form_edithermano=Formulario_hermano_candidato()
    form_hijo=Formulario_hijo_candidato
    form_edithijo=Formulario_hijo_candidato
    form_referencia=Formulario_referencia()
    form_editreferencia=Formulario_referencia()
    form_experiencia=Formulario_experiencia()
    form_editexperiencia=Formulario_experiencia()
    form_estudio=Formulario_estudios()
    form_editestudio=Formulario_estudios()
    
    form_estudiootro=Formulario_estudiosotros()
    form_editestudiootro=Formulario_estudiosotros()
        
    
    experiencias = Experiencia.objects.filter(candidato=candidato)
    referencias = Referencia.objects.filter(candidato=candidato)
    hijos = Hijo_candidato.objects.filter(candidato=candidato)
    hermanos = Hermano_candidato.objects.filter(candidato=candidato)
    idiomas = Idioma_candidato.objects.filter(candidato=candidato)
    estudios = Estudios_pro.objects.filter(candidato=candidato)
    estudiosotros=Estudios_otros.objects.filter(candidato=candidato)
        
    form_idioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"
    form_editidioma.fields['idioma'].widget.attrs['style'] ="text-transform: uppercase;"

        

    form_candidato.fields['edad'].widget.attrs['readonly'] ='readonly'
    form_candidato.fields['piso'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['depto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_que'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_donde'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_horario'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['estudia_termino'].widget.attrs['disabled'] ='disabled'

        
        
    form_candidato.fields['primaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['primaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['secundaria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['secundaria_documento'].widget.attrs['disabled'] ='disabled'
        
    form_candidato.fields['preparatoria_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['preparatoria_documento'].widget.attrs['disabled'] ='disabled'
      
    form_candidato.fields['tecnica_annios'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_inicio'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_termino'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['tecnica_documento'].widget.attrs['disabled'] ='disabled'

    

    form_candidato.fields['pago_infonavit'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_marca'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['auto_modelo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['seguro_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['afianzado_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['sindicato_cargo'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_monto'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['ingreso_fuente'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_nombre'].widget.attrs['disabled'] ='disabled'
    form_candidato.fields['conocido_depto'].widget.attrs['disabled'] ='disabled'

    form_experiencia.fields['experiencia_supervision_num'].widget.attrs['disabled'] ='disabled'
    

    return render(request, template_name, locals(),)

def lstcol_idioma(request):#buscar idioma
    #print(request.GET.get('id'))
    id=request.GET.get('id')
    
    idioma= Idioma_candidato.objects.get(pk=id)
    data = {
        'status':"OK",
        'id': idioma.id,
        'idioma': idioma.idioma,
        'porcentaje': idioma.idioma_porcentaje,
    }

    return JsonResponse(data)

def delcol_estudio(request):#Eliminar estudio
    data = dict()
    id=request.GET.get('id')
    estudio= Estudios_pro.objects.get(pk=id)
    candidato=estudio.candidato
    estudio.delete()
    data['form_is_valid'] = True
    estudios = Estudios_pro.objects.filter(candidato=candidato)
    data['html_estudios_lista'] = render_to_string('candidatos/estudios_lista.html', {
        'estudios': estudios
    })

    return JsonResponse(data)

def lstcol_estudio(request):#buscar estudio
    #print(request.GET.get('id'))
    id=request.GET.get('id')
    #id=13
    estudio= Estudios_pro.objects.get(pk=id)
    data = {
        'status':"OK",
        'id': estudio.id,
        'tipo': estudio.estudios_tipo,
        'escuela': estudio.estudios_escuela,
        'nombre': estudio.estudios_nombre,
        'annios': estudio.estudios_annios,
        'inicio': estudio.estudios_inicio,
        'termino': estudio.estudios_termino,
        'documento': estudio.estudios_documento,
        'tesis': estudio.estudios_tesis,
        'forma': estudio.estudios_forma,
        'cedula': estudio.estudios_cedula,
    }

    return JsonResponse(data)

def delcol_otroestudio(request):#Eliminar estudiootro
    data = dict()
    id=request.GET.get('id')
    estudio= Estudios_otros.objects.get(pk=id)
    candidato=estudio.candidato
    estudio.delete()
    data['form_is_valid'] = True
    estudiosotros = Estudios_otros.objects.filter(candidato=candidato)
    data['html_estudiosotros_lista'] = render_to_string('candidatos/estudiosotros_lista.html', {
        'estudiosotros': estudiosotros
    })

    return JsonResponse(data)

def lstcol_otroestudio(request):#buscar estudiootro
    #print(request.GET.get('id'))
    id=request.GET.get('id')
    
    estudio= Estudios_otros.objects.get(pk=id)
    data = {
        'status':"OK",
        'id': estudio.id,
        'tipo': estudio.estudios_tipo,
        'nombre': estudio.estudios_nombre,
        'inicio': estudio.estudios_inicio,
        'termino': estudio.estudios_termino,
        'documento': estudio.estudios_documento,
    }

    return JsonResponse(data)

def delcol_hijo(request):#borrar hijo
    data = dict()
    id=request.GET.get('id')
    hijo= Hijo_candidato.objects.get(pk=id)
    candidato=hijo.candidato
    hijo.delete()
    data['form_is_valid'] = True
    hijos = Hijo_candidato.objects.filter(candidato=candidato)
    data['html_hijos_lista'] = render_to_string('candidatos/hijos_lista.html', {
        'hijos': hijos
    })
    return JsonResponse(data)

def lstcol_hijo(request):#buscar hijo
    id=request.GET.get('id')
    hijo= Hijo_candidato.objects.get(pk=id)

    data = {
        'status':"OK",
        'id': hijo.id,
        'nombre': hijo.hijo_nombre,
        'paterno': hijo.hijo_apellido_paterno,
        'materno': hijo.hijo_apellido_materno,
        'edad': hijo.hijo_edad,
        'ocupacion': hijo.hijo_ocupacion,
        'lugar': hijo.hijo_lugar_ocupacion,
        'domicilio': hijo.hijo_domicilio,
        'telefono': hijo.hijo_tel,
    }

    return JsonResponse(data)

def delcol_hermano(request):#borrar hermano
    data = dict()
    id=request.GET.get('id')
    hermano= Hermano_candidato.objects.get(pk=id)
    candidato=hermano.candidato
    hermano.delete()
    data['form_is_valid'] = True
    hermanos = Hermano_candidato.objects.filter(candidato=candidato)
    data['html_hermanos_lista'] = render_to_string('candidatos/hermanos_lista.html', {
        'hermanos': hermanos
    })
    return JsonResponse(data)

def lstcol_hermano(request):#buscar hermano
    id=request.GET.get('id')
    hermano= Hermano_candidato.objects.get(pk=id)
    
    data = {
        'status':"OK",
        'id': hermano.id,
        'nombre': hermano.hermano_nombre,
        'paterno': hermano.hermano_apellido_paterno,
        'materno': hermano.hermano_apellido_materno,
        'edad': hermano.hermano_edad,
        'ocupacion': hermano.hermano_ocupacion,
        'lugar': hermano.hermano_lugar_ocupacion,
        'domicilio': hermano.hermano_domicilio,
        'telefono': hermano.hermano_tel,
    }

    return JsonResponse(data)

def delcol_experiencia(request):#borrar experiencia
    data = dict()
    id=request.GET.get('id')
    exp= Experiencia.objects.get(pk=id)
    candidato=exp.candidato
    exp.delete()
    data['form_is_valid'] = True
    experiencias = Experiencia.objects.filter(candidato=candidato)
    data['html_experiencias_lista'] = render_to_string('candidatos/experiencias_lista.html', {
        'experiencias': experiencias
    })
    return JsonResponse(data)

def lstcol_experiencia(request):#buscar hermano
    try:
       # print(request.GET.get('id'))
        id=request.GET.get('id')
        exp= Experiencia.objects.get(pk=id)
    except Exception as e:
        print(e)
    
    
    try:
        data = {
        'status':"OK",
        'id': exp.id,
        'nombre': exp.empresa_nombre,
        'direccion': exp.empresa_direccion,
        'telefono': exp.empresa_tel,
        'giro': exp.empresa_giro,
        'jefe': exp.empresa_nombre_jefe,
        'puesto': exp.empresa_jefe_puesto,
        'ingreso': exp.empresa_fecha_ingreso,
        'salario_inicial': exp.empresa_salario_inicio,
        'separacion': exp.empresa_fecha_separacion,
        'salario_final': exp.empresa_salario_final,
        'puesto_ultimo': exp.empresa_puesto_ultimo,
        'puesto_ultimo_tiempo': exp.empresa_puesto_ultimo_tiempo,
        'depto_ultimo': exp.empresa_puesto_ultimo_depto,
        'puesto_anterior': exp.empresa_puesto_anterior,
        'puesto_anterior_tiempo': exp.empresa_puesto_anterior_tiempo,
        'depto_anterior': exp.empresa_puesto_anterior_depto,
        'exp_supervision': exp.experiencia_supervision,
        'num_supervision': exp.experiencia_supervision_num,
        'motivo': exp.separacion_motivo,
    }
    
    except Exception as e:
        print(e)
    

    return JsonResponse(data)

def delcol_referencia(request):#borrar referencia
    data = dict()
    id=request.GET.get('id')
    ref= Referencia.objects.get(pk=id)
    candidato=ref.candidato
    ref.delete()
    data['form_is_valid'] = True
    referencias = Referencia.objects.filter(candidato=candidato)
    data['html_referencias_lista'] = render_to_string('candidatos/referencias_lista.html', {
        'referencias': referencias
    })
    return JsonResponse(data)

def lstcol_referencia(request):#buscar referencia
    try:
        id=request.GET.get('id')
        ref= Referencia.objects.get(pk=id)
    except Exception as e:
        print(e)
    try:
        data = {
        'status':"OK",
        'id': ref.id,
        'nombre': ref.referencia_nombre,
        'domicilio': ref.referencia_domicilio,
        'telefono': ref.referencia_tel,
        'ocupacion': ref.referencia_ocupacion,
        'annios': ref.referencia_annios_conocer,
    }
    
    except Exception as e:
        print(e)

    return JsonResponse(data)

def delcol_idioma(request):#elimina idioma
    data = dict()
    id=request.GET.get('id')
    
    idioma= Idioma_candidato.objects.get(pk=id)
    candidato=idioma.candidato
    idioma.delete()
    data['form_is_valid'] = True
    idiomas = Idioma_candidato.objects.filter(candidato=candidato)
    data['html_idiomas_lista'] = render_to_string('candidatos/idiomas_lista.html', {
        'idiomas': idiomas
    })
    return JsonResponse(data)

def col_secuno(request):#agregar secuno
    if not request.POST._mutable:
        request.POST._mutable = True
    request.POST['fecha_nac']=mod_fecha(request.POST.get('fecha_nac'))
    form = CandidatoSecunoForm(request.POST)
    response=""
    
    if form.is_valid():
        id_n = request.POST.get('id_n')
        con_id = request.POST.get('con_id')
        if con_id:
            id_n=con_id

        if id_n:
            candidato= Candidato.objects.get(id=id_n)
            
            candidato.fuente_recluta=form.cleaned_data['fuente_recluta']
            candidato.puesto_solicitado=form.cleaned_data['puesto_solicitado']
            candidato.sueldo_deseado=form.cleaned_data['sueldo_deseado']
            candidato.puesto_solicitado=form.cleaned_data['puesto_solicitado']
            candidato.nombre=form.cleaned_data['nombre']
            candidato.apellido_paterno=form.cleaned_data['apellido_paterno']
            candidato.apellido_materno=form.cleaned_data['apellido_materno']
            candidato.sexo=form.cleaned_data['sexo']
            candidato.estado_civil=form.cleaned_data['estado_civil']
            candidato.edad=form.cleaned_data['edad']
            candidato.fecha_nac=form.cleaned_data['fecha_nac']
            candidato.lugar_nac=form.cleaned_data['lugar_nac']
            candidato.pais_nacimiento=form.cleaned_data['pais_nacimiento']
            candidato.tel=form.cleaned_data['tel']
            candidato.cel=form.cleaned_data['cel']
            candidato.tipo=form.cleaned_data['tipo']
            candidato.calle=form.cleaned_data['calle']
            candidato.num_ext=form.cleaned_data['num_ext']
            candidato.num_int=form.cleaned_data['num_int']
            candidato.calle_uno=form.cleaned_data['calle_uno']
            candidato.calle_dos=form.cleaned_data['calle_dos']
            candidato.piso=form.cleaned_data['piso']
            candidato.depto=form.cleaned_data['depto']
            candidato.cp=form.cleaned_data['cp']
            candidato.colonia=form.cleaned_data['colonia']
            candidato.esdo=form.cleaned_data['esdo']
            candidato.pais_direc=form.cleaned_data['pais_direc']
            candidato.referencia=form.cleaned_data['referencia']
            candidato.trayectoria_de_casa=form.cleaned_data['trayectoria_de_casa']
            #candidato.email_personal=request.POST.get('email_personal')
            candidato.email_personal=form.cleaned_data['email_personal']
            candidato.rfc=form.cleaned_data['rfc']
            candidato.curp=form.cleaned_data['curp']
            candidato.imss=form.cleaned_data['imss']
            candidato.cartilla=form.cleaned_data['cartilla']
            candidato.tipo_licencia=form.cleaned_data['tipo_licencia']
            candidato.licencia=form.cleaned_data['licencia']
            
            #candidato.save()
            data = {
                'result': 'OK',
                'id':id_n
            }
            response=data
        else:
            print("edad:")
            print(form.cleaned_data['edad'])
            print(request.POST)
            #candidato= Candidato.objects.get(id=cand_id)
            #post=form.save()
            #print(post.id)
            data = {
                    'result': 'OK',
                    #'id':post.id
                    'id':13
            }
            response=data

    else:
        #pas_err=form.errors['password2']
        print(form.errors)
        #data = {
        #    'result': 'datos',
        #    'message': pas_err
        #}
        data = {
                'result': 'errores'
            }
        response=data
    return JsonResponse(response)

def col_secdos(request):#agregar secdos
    #tomar Id del input oculto
    cand_id = request.POST.get('candId2')
    print(cand_id)
    #cand_id=13
    if not request.POST._mutable:
        request.POST._mutable = True
    
    
    request.POST['estudia_termino']=mod_fecha(request.POST.get('estudia_termino'))

    #print(request.POST)
    form = CandidatoSecdosForm(request.POST)
    #print(form)
    if form.is_valid():
        try:
            #buscar candidato en base de datos
            candidato= Candidato.objects.get(id=cand_id)
            #candidato= Candidato.objects.get(id=37)
            
            ##asignar valores
            candidato.primaria=form.cleaned_data['primaria']
            candidato.primaria_annios=form.cleaned_data['primaria_annios']
            candidato.primaria_inicio=form.cleaned_data['primaria_inicio']
            candidato.primaria_termino=form.cleaned_data['primaria_termino']
            candidato.primaria_documento=form.cleaned_data['primaria_documento']
            candidato.secundaria=form.cleaned_data['secundaria']
            candidato.secundaria_annios=form.cleaned_data['secundaria_annios']
            candidato.secundaria_inicio=form.cleaned_data['secundaria_inicio']
            candidato.secundaria_termino=form.cleaned_data['secundaria_termino']
            candidato.secundaria_documento=form.cleaned_data['secundaria_documento']
            candidato.preparatoria=form.cleaned_data['preparatoria']
            candidato.preparatoria_annios=form.cleaned_data['preparatoria_annios']
            candidato.preparatoria_inicio=form.cleaned_data['preparatoria_inicio']
            candidato.preparatoria_termino=form.cleaned_data['preparatoria_termino']
            candidato.preparatoria_documento=form.cleaned_data['preparatoria_documento']
            candidato.tecnica=form.cleaned_data['tecnica']
            candidato.tecnica_annios=form.cleaned_data['tecnica_annios']
            candidato.tecnica_inicio=form.cleaned_data['tecnica_inicio']
            candidato.tecnica_termino=form.cleaned_data['tecnica_termino']
            candidato.tecnica_documento=form.cleaned_data['tecnica_documento']
            
            candidato.estudia_actualmente=form.cleaned_data['estudia_actualmente']
            candidato.estudia_que=form.cleaned_data['estudia_que']
            candidato.estudia_donde=form.cleaned_data['estudia_donde']
            candidato.estudia_horario=form.cleaned_data['estudia_horario']
            candidato.estudia_termino=form.cleaned_data['estudia_termino']
            candidato.maquinas_equipos=form.cleaned_data['maquinas_equipos']
            
            candidato.save()
            
            data = {
                'result': 'OK'
            }
            response=data
        except Exception as e:
            print(e)
            data = {
                'result': 'errores'
            }
            response=data
    else:
        print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data

    return JsonResponse(response)

def col_sectres(request):#agregar sectres
    #tomar Id del input oculto
    cand_id = request.POST.get('candId3')
    #cand_id=13
    form = CandidatoSectresForm(request.POST)
    if form.is_valid():
        try:
            #buscar candidato en base de datos
            candidato= Candidato.objects.get(id=cand_id)
            ##asignar valores
            candidato.padre_nombre=form.cleaned_data['padre_nombre']
            candidato.padre_apellido_paterno=form.cleaned_data['padre_apellido_paterno']
            candidato.padre_apellido_materno=form.cleaned_data['padre_apellido_materno']
            candidato.padre_edad=form.cleaned_data['padre_edad']
            candidato.padre_ocupacion=form.cleaned_data['padre_ocupacion']
            candidato.padre_lugar_trabajo=form.cleaned_data['padre_lugar_trabajo']
            candidato.padre_domicilio=form.cleaned_data['padre_domicilio']
            candidato.padre_tel=form.cleaned_data['padre_tel']
            candidato.padre_vive=form.cleaned_data['padre_vive']

            candidato.madre_nombre=form.cleaned_data['madre_nombre']
            candidato.madre_apellido_paterno=form.cleaned_data['madre_apellido_paterno']
            candidato.madre_apellido_materno=form.cleaned_data['madre_apellido_materno']
            candidato.madre_edad=form.cleaned_data['madre_edad']
            candidato.madre_ocupacion=form.cleaned_data['madre_ocupacion']
            candidato.madre_lugar_trabajo=form.cleaned_data['madre_lugar_trabajo']
            candidato.madre_domicilio=form.cleaned_data['madre_domicilio']
            candidato.madre_tel=form.cleaned_data['madre_tel']
            candidato.madre_vive=form.cleaned_data['madre_vive']

            candidato.conyuge_nombre=form.cleaned_data['conyuge_nombre']
            candidato.conyuge_apellido_paterno=form.cleaned_data['conyuge_apellido_paterno']
            candidato.conyuge_apellido_materno=form.cleaned_data['conyuge_apellido_materno']
            candidato.conyuge_edad=form.cleaned_data['conyuge_edad']
            candidato.conyuge_ocupacion=form.cleaned_data['conyuge_ocupacion']
            candidato.conyuge_lugar_trabajo=form.cleaned_data['conyuge_lugar_trabajo']
            candidato.conyuge_domicilio=form.cleaned_data['conyuge_domicilio']
            candidato.conyuge_tel=form.cleaned_data['conyuge_tel']
            candidato.conyuge_vive=form.cleaned_data['conyuge_vive']
            candidato.save()
            
            data = {
                'result': 'OK'
            }
            response=data
        except Exception as e:
           # print(e)
            data = {
                'result': 'errores'
            }
            response=data
    else:
        print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data

    return JsonResponse(response)

def col_seccuatro(request):#agregar secdos
    #tomar Id del input oculto
    cand_id = request.POST.get('candId4')
    #cand_id =13
    if not request.POST._mutable:
        request.POST._mutable = True
    request.POST['fecha_disponible']=mod_fecha(request.POST.get('fecha_disponible'))
    form = CandidatoSeccuatroForm(request.POST)
    if form.is_valid():
        try:
            #buscar candidato en base de datos
            candidato= Candidato.objects.get(id=cand_id)
            ##asignar valores
            candidato.vivienda_propia=form.cleaned_data['vivienda_propia']
            candidato.credito_infonavit=form.cleaned_data['credito_infonavit']
            candidato.pago_infonavit=form.cleaned_data['pago_infonavit']
            candidato.auto_propio=form.cleaned_data['auto_propio']
            candidato.auto_marca=form.cleaned_data['auto_marca']
            candidato.auto_modelo=form.cleaned_data['auto_modelo']
            candidato.seguro_vida=form.cleaned_data['seguro_vida']
            candidato.seguro_monto=form.cleaned_data['seguro_monto']

            candidato.afianzado=form.cleaned_data['afianzado']
            candidato.afianzado_monto=form.cleaned_data['afianzado_monto']
            candidato.afiliado_sindicato=form.cleaned_data['afiliado_sindicato']
            candidato.sindicato_nombre=form.cleaned_data['sindicato_nombre']
            candidato.sindicato_cargo=form.cleaned_data['sindicato_cargo']
            candidato.tiempo_libre=form.cleaned_data['tiempo_libre']
            candidato.embarazo=form.cleaned_data['embarazo']
            candidato.religion=form.cleaned_data['religion']

            candidato.estado_salud=form.cleaned_data['estado_salud']
            candidato.fuma=form.cleaned_data['fuma']
            candidato.bebe=form.cleaned_data['bebe']
            candidato.tatuajes=form.cleaned_data['tatuajes']
            candidato.perforaciones=form.cleaned_data['perforaciones']
            candidato.disposicion_rolar=form.cleaned_data['disposicion_rolar']
            candidato.disposicion_viajar=form.cleaned_data['disposicion_viajar']
            candidato.ingreso_extra=form.cleaned_data['ingreso_extra']
            candidato.ingreso_monto=form.cleaned_data['ingreso_monto']
            candidato.ingreso_fuente=form.cleaned_data['ingreso_fuente']
            candidato.labora_conocido=form.cleaned_data['labora_conocido']
            candidato.conocido_nombre=form.cleaned_data['conocido_nombre']
            candidato.conocido_depto=form.cleaned_data['conocido_depto']
            candidato.fecha_disponible=form.cleaned_data['fecha_disponible']
            candidato.save()
            
            data = {
                'result': 'OK'
            }
            response=data
        except Exception as e:
           # print(e)
            data = {
                'result': 'errores'
            }
            response=data
    else:
        #print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data

    return JsonResponse(response)

def col_seccinco(request):#agregar experiencia sec cinco
    data = dict()
    if not request.POST._mutable:
        request.POST._mutable = True
    request.POST['empresa_fecha_ingreso']=mod_fecha(request.POST.get('empresa_fecha_ingreso'))
    request.POST['empresa_fecha_separacion']=mod_fecha(request.POST.get('empresa_fecha_separacion'))
    form = ExperienciaSeccincoForm(request.POST)
    cand_id = request.POST.get('candId5')
    #cand_id =13
    response=""
    if form.is_valid():
        try:
            candidato= Candidato.objects.get(id=cand_id)
            experiencia = form.save(commit=False)
            experiencia.candidato=candidato
            experiencia.save()
            data['form_is_valid'] = True
            experiencias = Experiencia.objects.filter(candidato=candidato)
            data['html_experiencias_lista'] = render_to_string('candidatos/experiencias_lista.html', {
                'experiencias': experiencias
            })
            response=data
        except Exception as e:
            #print(e)
            data = {
                'result': 'errores'
            }
            response=data
    else:
        #print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data
    return JsonResponse(response)

def col_secseis(request):#agregar referencia sec seis
    data = dict()
    form = ReferenciaSecseisForm(request.POST)
    cand_id = request.POST.get('candId6')
    #cand_id =13
    response=""
    if form.is_valid():
        try:
            candidato= Candidato.objects.get(id=cand_id)
            referencia = form.save(commit=False)
            referencia.candidato=candidato
            referencia.save()
            data['form_is_valid'] = True
            referencias = Referencia.objects.filter(candidato=candidato)
            data['html_referencias_lista'] = render_to_string('candidatos/referencias_lista.html', {
                'referencias': referencias
            })
            response=data
        except Exception as e:
            #print(e)
            data = {
                'result': 'errores'
            }
            response=data
    else:
        #print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data
    return JsonResponse(response)

def addcol_hijo(request):#agregar hijo sec tres
    data = dict()
    form = HijoSectresForm(request.POST)
    bnombre = request.POST.get('hijo_nombre')
    bpaterno = request.POST.get('hijo_apellido_paterno')
    bmaterno = request.POST.get('hijo_apellido_materno')
    
    cand_id = request.POST.get('candId3_hijo')
    #cand_id =13
    response=""
    if form.is_valid():
        candidato= Candidato.objects.get(id=cand_id)
        try:
            hij=Hijo_candidato.objects.get(candidato=candidato,hijo_nombre=bnombre,hijo_apellido_paterno=bpaterno,hijo_apellido_materno=bmaterno)
            if hij:
                data={
                    'form_is_valid' : False,
                    'error':'Hijo ya existe!'
                }
        except Exception as e:
            #print(e)
            hijo = form.save(commit=False)
            hijo.candidato=candidato
            hijo.save()
            data['form_is_valid'] = True
            hijos = Hijo_candidato.objects.filter(candidato=candidato)
            data['html_hijos_lista'] = render_to_string('candidatos/hijos_lista.html', {
                'hijos': hijos
            })
        response=data
    else:
       # print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data
              
    return JsonResponse(response)

def editcol_hijo(request):#editar hijo
    data = dict()
    id=request.POST.get('hijo_id')
    cand_id=request.POST.get('candId_edithijo')
    #cand_id=13
    bnombre = request.POST.get('hijo_nombre')
    candidato= Candidato.objects.get(id=cand_id)
    try:
        hij=Hijo_candidato.objects.filter(~Q(id=id),candidato=candidato,hijo_nombre=bnombre)
        if hij:
            data={
                'form_is_valid' : False,
                'error':'Hijo ya existe!'
            }
        else:
            hijo= Hijo_candidato.objects.get(pk=id)
            candidato=hijo.candidato
            hijo.hijo_nombre=bnombre
            hijo.hijo_edad=request.POST.get('hijo_edad')
            hijo.hijo_ocupacion=request.POST.get('hijo_ocupacion')
    
            hijo.save()
            data['form_is_valid'] = True
            hijos = Hijo_candidato.objects.filter(candidato=candidato)
            data['html_hijos_lista'] = render_to_string('candidatos/hijos_lista.html', {
                  'hijos': hijos
            })
    except Exception as e:
        print(e)
        
    
    return JsonResponse(data)

def addcol_hermano(request):#agregar hermano sec tres
    data = dict()
    form = HermanoSectresForm(request.POST)
    bnombre = request.POST.get('hermano_nombre')
    bpaterno = request.POST.get('hermano_apellido_paterno')
    bmaterno = request.POST.get('hermano_apellido_materno')
    cand_id = request.POST.get('candId3_hermano')
    #cand_id =13
    response=""
    if form.is_valid():
        candidato= Candidato.objects.get(id=cand_id)
        try:
            her=Hermano_candidato.objects.get(candidato=candidato,hermano_nombre=bnombre,hermano_apellido_paterno=bpaterno,hermano_apellido_materno=bmaterno)
            if her:
                data={
                    'form_is_valid' : False,
                    'error':'Hermano ya existe!'
                }
        except Exception as e:            
            hermano = form.save(commit=False)
            hermano.candidato=candidato
            hermano.save()
            data['form_is_valid'] = True
            hermanos = Hermano_candidato.objects.filter(candidato=candidato)
            data['html_hermanos_lista'] = render_to_string('candidatos/hermanos_lista.html', {
                'hermanos': hermanos
            })
        response=data
    else:
        #print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data
    return JsonResponse(response)

def editcol_hermano(request):#editar hermano
    data = dict()
    id=request.POST.get('hermano_id')
    cand_id = request.POST.get('candId_edithermano')
    #cand_id =13
    candidato= Candidato.objects.get(id=cand_id)
    bnombre = request.POST.get('hermano_nombre')
    try:
        her=Hermano_candidato.objects.filter(~Q(id=id),candidato=candidato,hermano_nombre=bnombre)
        if her:
            data={
                'form_is_valid' : False,
                'error':'Hermano ya existe!'
            }
        else:
            hermano= Hermano_candidato.objects.get(pk=id)
            candidato=hermano.candidato
            hermano.hermano_nombre=request.POST.get('hermano_nombre')
            hermano.hermano_edad=request.POST.get('hermano_edad')
            hermano.hermano_ocupacion=request.POST.get('hermano_ocupacion')
    
            hermano.save()
            data['form_is_valid'] = True
            hermanos = Hermano_candidato.objects.filter(candidato=candidato)
            data['html_hermanos_lista'] = render_to_string('candidatos/hermanos_lista.html', {
                'hermanos': hermanos
            })

    except Exception as e:
        print(e)
       
    
    return JsonResponse(data)

def addcol_idioma(request):#agregar o modificar idioma
    data = dict()
    if not request.POST._mutable:
        request.POST._mutable = True
    form = IdiomaSecdosForm(request.POST)
    request.POST['idioma']=request.POST.get('idioma').upper()
    bidioma = request.POST.get('idioma')
    cand_id = request.POST.get('candId2_idi')
    #cand_id =13
    response=""
    if form.is_valid():
        id=request.POST.get('idi_id')
        candidato= Candidato.objects.get(id=cand_id)
        if (id in [None,'']):#nuevo idioma
            try:
                idi=Idioma_candidato.objects.get(candidato=candidato,idioma=bidioma)
                if idi:
                    data={
                        'form_is_valid' : False,
                        'error':'Idioma ya existe!'
                    }
            except Exception as e:
                idi = form.save(commit=False)
                idi.candidato=candidato
                idi.save()
                data['form_is_valid'] = True
                idiomas = Idioma_candidato.objects.filter(candidato=candidato)
                data['html_idiomas_lista'] = render_to_string('candidatos/idiomas_lista.html', {
                    'idiomas': idiomas
                },request=request)
            response=data     
        else:#actualizar idioma
            try:
                
                idioma= Idioma_candidato.objects.get(pk=id)
                #idioma = form.save(commit=False)
                idioma.idioma=form.cleaned_data['idioma']
                idioma.idioma_porcentaje=form.cleaned_data['idioma_porcentaje']
                idioma.save()
                data['form_is_valid'] = True
                idiomas = Idioma_candidato.objects.filter(candidato=candidato)
                data['html_idiomas_lista'] = render_to_string('candidatos/idiomas_lista.html', {
                    'idiomas': idiomas
                },request=request)
                response=data
            except Exception as e:
                print(e)
                data = {
                    'result': 'errores'
                }

    else:
        print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data
              
    return JsonResponse(response)

def editcol_idioma(request):#editar idioma
    data = dict()
    if not request.POST._mutable:
        request.POST._mutable = True
    id=request.POST.get('idi_id')
    request.POST['idioma']=request.POST.get('idioma').upper()
    f_idioma=request.POST.get('idioma')
    #print(f_idioma)
    f_porcentaje=request.POST.get('idioma_porcentaje')
    cand_id = request.POST.get('candId_editidioma')
    #cand_id =13
    candidato= Candidato.objects.get(id=cand_id)
    #investiga si existe el nombre
    try:
        
        idi=Idioma_candidato.objects.filter(~Q(id=id),candidato=candidato,idioma=f_idioma)
        if idi:
            data={
                    'form_is_valid' : False,
                    'error':'Idioma ya existe!'
                }
        else:
            idioma= Idioma_candidato.objects.get(pk=id)
            candidato=idioma.candidato

            idioma.idioma=f_idioma.upper()
            idioma.idioma_porcentaje=f_porcentaje
            idioma.save()
            data['form_is_valid'] = True
            idiomas = Idioma_candidato.objects.filter(candidato=candidato)
            data['html_idiomas_lista'] = render_to_string('candidatos/idiomas_lista.html', {
               'idiomas': idiomas
            })
       # print("encontro")
    except Exception as e:
        print(e)
        
    
    
    return JsonResponse(data)

def addcol_otroestudio(request):#agregar estudiootro sec dos
    data = dict()
    form = EstudioOtroSecdosForm(request.POST)
    cand_id = request.POST.get('candId2_estotro')
   # cand_id =13
    response=""
    if form.is_valid():
        candidato= Candidato.objects.get(id=cand_id)
        estudio = form.save(commit=False)
        estudio.candidato=candidato
        estudio.save()
        data['form_is_valid'] = True
        estudiosotros = Estudios_otros.objects.filter(candidato=candidato)
        data['html_estudiosotros_lista'] = render_to_string('candidatos/estudiosotros_lista.html', {
            'estudiosotros': estudiosotros
        })
        response=data
    else:
        print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data
              
    return JsonResponse(response)

def editcol_otroestudio(request):#editar estudio
    data = dict()
    id=request.POST.get('estotro_id')

    try:
        estudio= Estudios_otros.objects.get(pk=id)
        candidato=estudio.candidato
        estudio.estudios_tipo=request.POST.get('estudios_tipo')
        estudio.estudios_nombre=request.POST.get('estudios_nombre')
        estudio.estudios_inicio=request.POST.get('estudios_inicio')
        estudio.estudios_termino=request.POST.get('estudios_termino')
        estudio.estudios_documento=request.POST.get('estudios_documento')
    
        estudio.save()
        data['form_is_valid'] = True
        estudiosotros = Estudios_otros.objects.filter(candidato=candidato)
        data['html_estudiosotros_lista'] = render_to_string('candidatos/estudiosotros_lista.html', {
            'estudiosotros': estudiosotros
        })
        response=data
    except Exception as e:
        print(e)
        data = {
            'result': 'errores'
        }
        response=data
    
    
    return JsonResponse(data)

def addcol_estudio(request):#agregar estudio sec dos
    data = dict()
    form = EstudioSecdosForm(request.POST)
    cand_id = request.POST.get('candId2_est')
    #cand_id =13
    response=""
    if form.is_valid():
        candidato= Candidato.objects.get(id=cand_id)
        estudio = form.save(commit=False)
        estudio.candidato=candidato
        estudio.save()
        data['form_is_valid'] = True
        estudios = Estudios_pro.objects.filter(candidato=candidato)
        data['html_estudios_lista'] = render_to_string('colaboradores/estudios_lista.html', {
            'estudios': estudios
        })
        response=data
    else:
        print(form.errors)
        data = {
                'result': 'errores'
        }
        response=data
              
    return JsonResponse(response)

def editcol_estudio(request):#editar estudio
    data = dict()
    id=request.POST.get('est_id')

    try:
        estudio= Estudios_pro.objects.get(pk=id)
        candidato=estudio.candidato
        estudio.estudios_tipo=request.POST.get('estudios_tipo')
        estudio.estudios_escuela=request.POST.get('estudios_escuela')
        estudio.estudios_nombre=request.POST.get('estudios_nombre')
        estudio.estudios_annios=request.POST.get('estudios_annios')
        estudio.estudios_inicio=request.POST.get('estudios_inicio')
        estudio.estudios_termino=request.POST.get('estudios_termino')
        estudio.estudios_documento=request.POST.get('estudios_documento')
        estudio.estudios_tesis=request.POST.get('estudios_tesis')
        estudio.estudios_forma=request.POST.get('estudios_forma')
        estudio.estudios_cedula=request.POST.get('estudios_cedula')
    
        estudio.save()
        data['form_is_valid'] = True
        estudios = Estudios_pro.objects.filter(candidato=candidato)
        data['html_estudios_lista'] = render_to_string('candidatos/estudios_lista.html', {
            'estudios': estudios
        })
        response=data
    except Exception as e:
        print(e)
        data = {
            'result': 'errores'
        }
        response=data
    
    
    return JsonResponse(data)

def editcol_experiencia(request):#editar experiencia
    data = dict()
    #form = IdiomaSecdosForm(request.POST)
    if not request.POST._mutable:
        request.POST._mutable = True
    request.POST['empresa_fecha_ingreso']=mod_fecha(request.POST.get('empresa_fecha_ingreso'))
    request.POST['empresa_fecha_separacion']=mod_fecha(request.POST.get('empresa_fecha_separacion'))
    id=request.POST.get('exp_id')
    exp= Experiencia.objects.get(pk=id)
    candidato=exp.candidato
    exp.empresa_nombre=request.POST.get('empresa_nombre')
    exp.empresa_direccion=request.POST.get('empresa_direccion')
    exp.empresa_tel=request.POST.get('empresa_tel')
    exp.empresa_giro=request.POST.get('empresa_giro')
    exp.empresa_nombre_jefe=request.POST.get('empresa_nombre_jefe')
    exp.empresa_jefe_puesto=request.POST.get('empresa_jefe_puesto')
    exp.empresa_fecha_ingreso=request.POST.get('empresa_fecha_ingreso')
    exp.empresa_salario_inicio=request.POST.get('empresa_salario_inicio')
    exp.empresa_fecha_separacion=request.POST.get('empresa_fecha_separacion')
    exp.empresa_salario_final=request.POST.get('empresa_salario_final')
    exp.empresa_puesto_ultimo=request.POST.get('empresa_puesto_ultimo')
    exp.empresa_puesto_ultimo_tiempo=request.POST.get('empresa_puesto_ultimo_tiempo')
    exp.empresa_puesto_ultimo_depto=request.POST.get('empresa_puesto_ultimo_depto')
    exp.empresa_puesto_anterior=request.POST.get('empresa_puesto_anterior')
    exp.empresa_puesto_anterior_tiempo=request.POST.get('empresa_puesto_anterior_tiempo')
    exp.empresa_puesto_anterior_depto=request.POST.get('empresa_puesto_anterior_depto')
    exp.experiencia_supervision=request.POST.get('experiencia_supervision')
    exp.experiencia_supervision_num=request.POST.get('experiencia_supervision_num')
    exp.separacion_motivo=request.POST.get('separacion_motivo')

    
    exp.save()
    data['form_is_valid'] = True
    experiencias = Experiencia.objects.filter(candidato=candidato)
    data['html_experiencias_lista'] = render_to_string('candidatos/experiencias_lista.html', {
        'experiencias': experiencias
    })
    response=data
    
    return JsonResponse(data)

def editcol_referencia(request):#editar referencia
    data = dict()
    #form = IdiomaSecdosForm(request.POST)
    
    id=request.POST.get('ref_id')
    ref= Referencia.objects.get(pk=id)
    candidato=ref.candidato
    ref.referencia_nombre=request.POST.get('referencia_nombre')
    ref.referencia_domicilio=request.POST.get('referencia_domicilio')
    ref.referencia_tel=request.POST.get('referencia_tel')
    ref.referencia_ocupacion=request.POST.get('referencia_ocupacion')
    ref.referencia_annios_conocer=request.POST.get('referencia_annios_conocer')
        
    ref.save()
    data['form_is_valid'] = True
    referencias = Referencia.objects.filter(candidato=candidato)
    data['html_referencias_lista'] = render_to_string('candidatos/referencias_lista.html', {
        'referencias': referencias
    })
    response=data
    
    return JsonResponse(data)



####